import argparse
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple


def _require_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise SystemExit(
            "openpyxl is required to read .xlsx files.\n"
            "Tip: activate the repo venv and run:\n"
            "  source .venv-worker/bin/activate && python scripts/convert_csv_v6.py\n"
        ) from e
    return load_workbook


# ==========================================
# 1. 配置：输入文件与输出目标
# ==========================================
DEFAULT_XLSX = os.path.expanduser("~/Desktop/skincare_ingredients_and_comparisons_updated.xlsx")
OUTPUT_FILE = "batch_final.json"

SHEETS = [
    "Ingredients_Collected",
    "Comparison_Cleansers",
    "Comparison_TonersAcids",
    "Comparison_Serums",
    "Comparison_CreamsSPF",
]

# ==========================================
# 2. 价格神谕 (Price Oracle)
# ==========================================
PRICE_ORACLE = {
    # --- Luxury ---
    "La Mer": (380, 2800),
    "Helena Rubinstein": (450, 3680),
    "SkinCeuticals": (160, 1400),
    "Estee Lauder": (115, 960),
    "Lancome": (130, 1100),
    "Drunk Elephant": (80, 600),
    "Murad": (92, 698),
    "Clarins": (90, 750),
    "Fresh": (45, 450),
    # --- Mid ---
    "Kiehl's": (38, 330),
    "EltaMD": (41, 298),
    "Clinique": (32, 250),
    "La Roche-Posay": (25, 180),
    "Vichy": (30, 220),
    "Avene": (28, 198),
    "Bioderma": (18, 158),
    "Paula's Choice": (35, 320),
    # --- Budget ---
    "CeraVe": (16, 128),
    "Neutrogena": (20, 140),
    "The Ordinary": (12, 89),
    "Stridex": (8, 55),
    "PanOxyl": (12, 90),
    "Vanicream": (15, 110),
    "Inkey List": (12, 99),
    # --- Asian ---
    "Winona": (45, 198),
    "Wei Nuo Na": (45, 198),
    "Freeplus": (30, 150),
    "Curél": (25, 138),
    "Hada Labo": (18, 99),
    "Anessa": (45, 248),
    "Dr. Wu": (40, 240),
    "Skin Aqua": (15, 79),
    "Canmake": (12, 85),
    "COSRX": (25, 160),
    "Rohto": (15, 88),
}

# ==========================================
# 3. 品牌与地域规则 (Brand & Region Logic)
# ==========================================
BRAND_RULES = {
    "EltaMD": ["US"],
    "CeraVe": ["Global"],
    "Paula's Choice": ["Global"],
    "Stridex": ["US", "CN"],
    "Murad": ["US", "Global"],
    "PanOxyl": ["US"],
    "Vanicream": ["US"],
    "Neutrogena": ["Global"],
    "Drunk Elephant": ["US", "Global"],
    "Fresh": ["Global"],
    "SkinCeuticals": ["Global"],
    "Winona": ["CN"],
    "Wei Nuo Na": ["CN"],
    "Freeplus": ["CN", "Asia"],
    "Curél": ["CN", "Asia"],
    "Anessa": ["Asia"],
    "Hada Labo": ["Asia"],
    "Dr. Wu": ["Asia"],
    "Canmake": ["Asia"],
    "Skin Aqua": ["Asia"],
    "Bio-Essence": ["Asia"],
    "COSRX": ["Global"],
    "Rohto": ["Asia"],
    "La Roche-Posay": ["Global", "EU"],
    "Vichy": ["Global", "EU"],
    "Avene": ["Global", "EU"],
    "Bioderma": ["Global", "EU"],
    "Medik8": ["EU", "Global"],
    "Geek & Gorgeous": ["EU"],
    "Estee Lauder": ["Global"],
    "Lancome": ["Global"],
    "Kiehl's": ["Global"],
    "Clinique": ["Global"],
    "The Ordinary": ["Global"],
    "L'Oreal": ["Global"],
    "Helena Rubinstein": ["Global", "CN", "EU"],
}


def _norm(s: str) -> str:
    if s is None:
        return ""
    text = str(s)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text


def _norm_lower(s: str) -> str:
    return _norm(s).lower()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    return s


def clean_ingredients(text: Any) -> str:
    cleaned = clean_text(text)
    if not cleaned:
        return ""
    cleaned = cleaned.replace("Active Ingredients:", "").replace("Inactive Ingredients:", "")
    return cleaned.strip()


def normalize_product_key(name: str) -> str:
    s = _norm_lower(name).strip()
    if not s:
        return ""
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def estimate_prices(product_name: str, brand: str) -> Tuple[float, float]:
    """
    双币种价格估算
    """
    name_norm = _norm_lower(product_name)
    brand_norm = _norm_lower(brand)

    # 1. 精确匹配（按名字包含品牌关键词）
    for key, (usd, cny) in PRICE_ORACLE.items():
        if _norm_lower(key) in name_norm:
            return float(usd), float(cny)

    # 2. 品牌匹配
    for key, (usd, cny) in PRICE_ORACLE.items():
        if _norm_lower(key) == brand_norm:
            return float(usd), float(cny)

    # 3. 默认兜底
    return 35.0, 250.0


def _category_from_sheet(sheet_name: str, category_cell: str = "") -> str:
    sheet_lower = sheet_name.lower()
    cell_lower = _norm_lower(category_cell)

    if "cleanser" in sheet_lower or "cleanser" in cell_lower or "洁面" in category_cell:
        return "Cleanser"
    if "toner" in sheet_lower or "acid" in sheet_lower or "toner" in cell_lower or "acid" in cell_lower or "水" in category_cell:
        return "Toner/Acid"
    if "serum" in sheet_lower or "serum" in cell_lower or "精华" in category_cell:
        return "Serum"
    if "cream" in sheet_lower or "spf" in sheet_lower or "cream" in cell_lower or "spf" in cell_lower or "面霜" in category_cell:
        return "Cream/Sunscreen"
    return "Treatment"


def infer_metadata(product_name: str, sheet_name: str, category_cell: str = "") -> Tuple[str, List[str], str]:
    """
    推断品牌、地域、分类
    """
    clean_name = clean_text(product_name)
    brand = "Unknown"
    regions: List[str] = ["Global"]

    clean_norm = _norm_lower(clean_name)

    # 品牌匹配（优先长名字）
    for b_key in sorted(BRAND_RULES.keys(), key=len, reverse=True):
        if _norm_lower(b_key) in clean_norm:
            brand = b_key
            regions = list(BRAND_RULES[b_key])
            break

    if brand == "Unknown":
        # 常见缩写补丁
        if re.search(r"\blrp\b", clean_norm):
            brand = "La Roche-Posay"
            regions = list(BRAND_RULES.get(brand, ["Global"]))
        else:
            # fallback: first token
            brand = clean_name.split(" ")[0].strip() or "Unknown"
            # if this token matches a known brand after normalization
            for b_key in BRAND_RULES.keys():
                if _norm_lower(b_key) == _norm_lower(brand):
                    brand = b_key
                    regions = list(BRAND_RULES[b_key])
                    break

    category = _category_from_sheet(sheet_name, category_cell)
    return brand, regions, category


def _dedupe_join(existing: str, new: str) -> str:
    a = clean_text(existing)
    b = clean_text(new)
    if not b:
        return a
    if not a:
        return b
    parts: List[str] = []
    seen = set()
    for chunk in (a + " | " + b).split("|"):
        c = chunk.strip()
        if not c:
            continue
        key = _norm_lower(c)
        if key in seen:
            continue
        seen.add(key)
        parts.append(c)
    return " | ".join(parts)


def _add_region_from_notes(regions: List[str], notes: str) -> List[str]:
    if not notes:
        return regions
    t = _norm_lower(notes)
    if any(k in t for k in ["cn", "china", "国内", "国行", "淘宝", "天猫", "京东"]):
        regions.append("CN")
    if any(k in t for k in ["us", "usa", "sephora", "amazon", "美国"]):
        regions.append("US")
    if any(k in t for k in ["eu", "europe", "欧洲", "uk", "英国"]):
        regions.append("EU")
    if any(k in t for k in ["jp", "japan", "日本"]):
        regions.append("Asia")
    return regions


def _normalize_regions(regions: List[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for r in regions:
        v = str(r or "").strip()
        if not v:
            continue
        key = v.upper()
        if key in seen:
            continue
        seen.add(key)
        out.append("Global" if key == "GLOBAL" else v.upper() if key in {"CN", "US", "EU"} else v)
    if not out:
        return ["Global"]
    return out


def strip_brand_prefix(full_name: str, brand: str) -> str:
    name = clean_text(full_name)
    if not name:
        return ""
    brand_clean = clean_text(brand)
    if not brand_clean or brand_clean == "Unknown":
        return name
    name_norm = _norm_lower(name)
    brand_norm = _norm_lower(brand_clean)
    if name_norm.startswith(brand_norm):
        rest = name[len(brand_clean) :].strip()
        rest = rest.lstrip("-–—:").strip()
        return rest or name
    return name


@dataclass
class ProductAcc:
    product_key: str
    brand: str
    name: str
    ingredients: str
    category: str
    availability: List[str]
    price_usd: float
    price_cny: float
    expert_knowledge: Dict[str, str]


def _ensure_expert_obj() -> Dict[str, str]:
    return {"sensitivity_flags": "", "chemist_notes": "", "key_actives": ""}


def _merge_expert(dst: Dict[str, str], src: Dict[str, str]) -> Dict[str, str]:
    for k in ["sensitivity_flags", "chemist_notes", "key_actives"]:
        dst[k] = _dedupe_join(dst.get(k, ""), src.get(k, ""))
    return dst


def _sheet_rows(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        row: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            row[h] = values[idx] if idx < len(values) else None
        out.append(row)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert Aurora skincare xlsx (5 sheets) into batch_final.json with stable expert_knowledge.")
    parser.add_argument("--input", default=DEFAULT_XLSX, help=f"Path to xlsx (default: {DEFAULT_XLSX})")
    parser.add_argument("--output", default=OUTPUT_FILE, help=f"Output JSON file (default: {OUTPUT_FILE})")
    args = parser.parse_args()

    load_workbook = _require_openpyxl()

    if not os.path.exists(args.input):
        raise SystemExit(f"Input file not found: {args.input}")

    wb = load_workbook(args.input, read_only=True, data_only=True)

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Workbook missing sheets: {missing}. Found: {wb.sheetnames}")

    by_key: Dict[str, ProductAcc] = {}

    def upsert_product(
        *,
        sheet: str,
        product_full: str,
        ingredients: str,
        category_cell: str = "",
        region_notes: str = "",
        expert: Optional[Dict[str, str]] = None,
    ) -> None:
        if not product_full or len(product_full) < 2:
            return

        brand, regions, category = infer_metadata(product_full, sheet, category_cell)
        name = strip_brand_prefix(product_full, brand)

        # Region patch: Chinese INCI or explicit region notes => include CN.
        if re.search(r"[\u4e00-\u9fff]", ingredients):
            regions.append("CN")
        regions = _add_region_from_notes(regions, region_notes or "")
        regions = _normalize_regions(regions)

        price_usd, price_cny = estimate_prices(product_full, brand)

        key = normalize_product_key(product_full)
        if not key:
            return

        ek = _ensure_expert_obj()
        if expert:
            ek = _merge_expert(ek, expert)

        existing = by_key.get(key)
        if not existing:
            by_key[key] = ProductAcc(
                product_key=key,
                brand=brand,
                name=name,
                ingredients=ingredients,
                category=category,
                availability=regions,
                price_usd=price_usd,
                price_cny=price_cny,
                expert_knowledge=ek,
            )
            return

        # Merge fields
        if existing.brand == "Unknown" and brand != "Unknown":
            existing.brand = brand
        # Prefer a cleaned name without the brand prefix.
        if existing.name == existing.product_key or not existing.name:
            existing.name = name or existing.name
        # Prefer main-sheet ingredients if available, otherwise keep first non-empty.
        if ingredients and (not existing.ingredients or sheet == "Ingredients_Collected"):
            existing.ingredients = ingredients
        if category and (existing.category == "Treatment" or sheet == "Ingredients_Collected"):
            existing.category = category

        existing.availability = _normalize_regions(existing.availability + regions)
        # Prefer oracle prices if existing is default.
        if (existing.price_usd, existing.price_cny) == (35.0, 250.0) and (price_usd, price_cny) != (35.0, 250.0):
            existing.price_usd, existing.price_cny = price_usd, price_cny

        existing.expert_knowledge = _merge_expert(existing.expert_knowledge, ek)

    # 1) Main sheet: source of truth for ingredients + category.
    ws_main = wb["Ingredients_Collected"]
    for row in _sheet_rows(ws_main):
        product = clean_text(row.get("Product"))
        ingredients = clean_ingredients(row.get("Ingredients (as listed)"))
        category_cell = clean_text(row.get("Category"))
        variant_notes = clean_text(row.get("Variant / Region Notes"))
        notes = clean_text(row.get("Notes"))
        if not product or len(ingredients) < 5:
            continue
        upsert_product(
            sheet="Ingredients_Collected",
            product_full=product,
            ingredients=ingredients,
            category_cell=category_cell,
            region_notes=variant_notes,
            expert={"chemist_notes": notes} if notes else None,
        )

    # 2) Comparison sheets: enrich expert_knowledge buckets.
    sheet_specs = {
        "Comparison_Cleansers": {
            "chemist_notes": ["Comparison notes", "Best for (rule of thumb)"],
            "sensitivity_flags": ["Fragrance/EO notes"],
            "key_actives": [],
        },
        "Comparison_TonersAcids": {
            "chemist_notes": ["Comparison notes"],
            "sensitivity_flags": ["Irritant flags (rule of thumb)"],
            "key_actives": ["Strength / Key actives"],
        },
        "Comparison_Serums": {
            "chemist_notes": ["Comparison notes"],
            "sensitivity_flags": ["Sensitivity flags"],
            "key_actives": ["Key actives (typical)"],
        },
        "Comparison_CreamsSPF": {
            "chemist_notes": ["Comparison notes"],
            "sensitivity_flags": ["Sensitivity flags"],
            "key_actives": ["Key ingredients/filters (typical)"],
        },
    }

    for sheet in ["Comparison_Cleansers", "Comparison_TonersAcids", "Comparison_Serums", "Comparison_CreamsSPF"]:
        ws = wb[sheet]
        spec = sheet_specs[sheet]
        for row in _sheet_rows(ws):
            product = clean_text(row.get("Product"))
            ingredients = clean_ingredients(row.get("Ingredients (as listed)"))
            if not product or len(ingredients) < 5:
                continue

            expert = _ensure_expert_obj()

            def add_block(dst_key: str, labels: List[str], prefix: str = ""):
                for label in labels:
                    val = clean_text(row.get(label))
                    if not val:
                        continue
                    content = f"{prefix}{val}" if prefix else val
                    expert[dst_key] = _dedupe_join(expert.get(dst_key, ""), content)

            add_block("sensitivity_flags", spec["sensitivity_flags"])
            add_block("key_actives", spec["key_actives"])
            # For cleansers, keep "Best for" inside chemist_notes (helps routine fit).
            if sheet == "Comparison_Cleansers":
                add_block("chemist_notes", ["Best for (rule of thumb)"], prefix="Best for: ")
                add_block("chemist_notes", ["Comparison notes"], prefix="Comparison: ")
            else:
                add_block("chemist_notes", spec["chemist_notes"])

            upsert_product(
                sheet=sheet,
                product_full=product,
                ingredients=ingredients,
                expert=expert,
            )

    wb.close()

    items: List[Dict[str, Any]] = []
    for acc in by_key.values():
        items.append(
            {
                "brand": acc.brand,
                "name": acc.name,
                "price_usd": acc.price_usd,
                "price_cny": acc.price_cny,
                "ingredients": acc.ingredients,
                "category": acc.category,
                "availability": acc.availability,
                "expert_knowledge": {
                    "sensitivity_flags": acc.expert_knowledge.get("sensitivity_flags", ""),
                    "chemist_notes": acc.expert_knowledge.get("chemist_notes", ""),
                    "key_actives": acc.expert_knowledge.get("key_actives", ""),
                },
            }
        )

    # Deterministic output order for diffs
    items.sort(key=lambda x: (_norm_lower(x.get("brand", "")), _norm_lower(x.get("name", ""))))

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)

    print(f"✅ Generated '{args.output}' with {len(items)} SKUs from '{args.input}'.")
    print("   Next: python worker/ingest.py --input-json batch_final.json --social-provider llm --overwrite")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

