import argparse
import hashlib
import json
import os
import re
import time
import uuid
import unicodedata
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple

import psycopg2
import requests
from dotenv import load_dotenv
from psycopg2.extras import Json
from openpyxl import load_workbook


SYSTEM_PROMPT = """
You are the Aurora Vectorization Engine.
Analyze the skincare product ingredients and return a JSON object with scores (0-100).

Rules (high-level heuristics):
- Oil Control: High if Alcohol, Caffeine, Clay, BHA/Salicylic Acid.
- Soothing: High if Centella, Panthenol, Allantoin, Madecassoside.
- Anti-aging: High if Retinoids, Peptides, Vitamin C, Niacinamide.
- Barrier repair: High if Ceramides, Cholesterol, Fatty Acids, Panthenol.

Risk flags:
- Use this controlled vocabulary (only set when you are confident from the ingredient list):
  - 'alcohol_high' if 'Alcohol Denat' (denatured alcohol) appears in top 5.
  - 'strong_acid' if strong exfoliating acids (e.g., glycolic/lactic/salicylic) are present as key actives.
  - 'mild_acid' if milder acids (e.g., azelaic/mandelic/PHA) are present.
  - 'retinol_high' if retinoids are present (retinol/retinal/tretinoin/adapalene).
  - 'benzoyl_peroxide' if benzoyl peroxide is present.
  - 'fragrance' if Fragrance/Parfum or common fragrance allergens are present.
  - 'mint' if menthol/peppermint/camphor/eucalyptus are present.
  - 'fungal_acne' if Polysorbates are present.
  - 'high_irritation' ONLY if strong_acid/retinol_high/benzoyl_peroxide is present (leave-on "strong actives").

Output Format (JSON):
{
  "mechanism": {
    "oil_control": int,
    "anti_aging": int,
    "soothing": int,
    "barrier_repair": int
  },
  "risk_flags": [str],
  "experience_prediction": {
    "texture": str,
    "finish": "matte|dewy|natural"
  },
  "social_stats": {
    "red_score": int,
    "reddit_score": int,
    "burn_rate": number,
    "top_keywords": [str]
  }
}
""".strip()

SOCIAL_PROMPT = """
You are a Beauty Trend Analyst. Based on your internal training data (internet discussions from Reddit/SkincareAddiction, XiaoHongShu/RED, TikTok), estimate the social sentiment for this product.

Product: {brand} - {name}

Rules for Estimation:
1) RED Score (0-100): High if popular in Asia, whitening/texture focused. Penalty for "Fake Slip" (假滑).
2) Reddit Score (0-100): High if ingredient-focused, fragrance-free, matte finish.
3) Burn Rate (0.0 - 1.0): Estimate probability of irritation complaints (e.g., 0.15 for high acid/retinol, 0.01 for gentle cleansers). Use >0.30 only for extremely irritating formulas.
4) Keywords: Extract 3-5 typical user tags (e.g., "HolyGrail", "Stings", "Pilling").

Output JSON:
{
  "redScore": int,
  "redditScore": int,
  "burnRate": float,
  "topKeywords": [str]
}
""".strip()

DEFAULT_GEMINI_API_BASE_URL = "https://generativelanguage.googleapis.com/v1beta"
DEFAULT_EMBEDDING_DIM = 1536
ENV_TEMPLATE_RE = re.compile(r"\$\{\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*\}\}")

# Brand/product aliases (DB-backed) for better anchor resolution in chat.
#
# - Keys are normalized via `normalize_alias_text`.
# - Values are aliases users may type (CN/EN/nicknames).
BRAND_ALIAS_SYNONYMS: Dict[str, List[str]] = {
  "skinceuticals": ["修丽可", "杜克", "skinceuticals"],
  "larocheposay": ["理肤泉", "lrp", "la roche posay"],
  "curel": ["珂润", "curel", "curél"],
  "winona": ["薇诺娜", "winona", "winnona"],
  "freeplus": ["芙丽芳丝", "freeplus"],
  "cerave": ["cerave", "适乐肤"],
  "lamer": ["海蓝之谜", "la mer", "lamer"],
  "esteelauder": ["雅诗兰黛", "estee lauder", "estée lauder", "anr"],
  "paulaschoice": ["宝拉", "paula's choice", "paulas choice", "pc"],
  "theordinary": ["the ordinary", "ordinary", "to", "theordinary"],
  "avene": ["雅漾", "avène", "avene"],
  "bioderma": ["贝德玛", "bioderma"],
  "vichy": ["薇姿", "vichy"],
  "neutrogena": ["露得清", "neutrogena"],
  "hadalabo": ["肌研", "hada labo", "hadalabo"],
  "cosrx": ["cosrx"],
  "vanicream": ["vanicream"],
  "murad": ["murad", "慕拉得", "慕拉德"],
}

# Product nicknames keyed by normalized "brand+name".
PRODUCT_NICKNAMES: Dict[str, List[str]] = {
  "larocheposaycicaplastbaumeb5": ["b5修护霜", "b5面霜", "b5霜", "cicaplast b5", "baume b5"],
  "skinceuticalsceferulic": ["cef", "ce ferulic", "ceferulic", "修丽可ce"],
  "lamercrmedelamer": ["海蓝之谜面霜", "la mer 面霜", "creme de la mer", "crème de la mer"],
  "niveacreme": ["妮维雅蓝罐", "蓝罐", "nivea 蓝罐"],
  "paulaschoiceskinperfecting2bhaliquidexfoliant": ["宝拉2%水杨酸", "2% bha", "pc2bha", "水杨酸2%"],
}


def normalize_alias_text(text: str) -> str:
  """
  Normalize user-entered aliases while keeping CJK characters.

  - NFKC (full-width -> half-width)
  - lowercase
  - remove whitespace/punctuation
  - keep only [a-z0-9] + CJK range
  """
  raw = unicodedata.normalize("NFKC", str(text or ""))
  raw = raw.lower()
  raw = re.sub(r"\s+", "", raw)
  raw = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", raw)
  return raw


@dataclass(frozen=True)
class InputSku:
  brand: str
  name: str
  ingredients_text: str
  price_usd: float
  price_cny: float
  availability: Optional[List[str]] = None
  product_url: Optional[str] = None
  image_url: Optional[str] = None
  expert_knowledge: Optional[Dict[str, Any]] = None
  kb_snippets: Optional[List[Dict[str, Any]]] = None


@dataclass(frozen=True)
class KbSnippet:
  brand: str
  name: str
  source_sheet: str
  field: str
  content: str
  metadata: Dict[str, Any]


def _require_env(name: str) -> str:
  value = (os.getenv(name) or "").strip()
  if not value:
    raise RuntimeError(f"Missing required env var: {name}")
  return value


def _require_any_env(names: List[str]) -> str:
  for name in names:
    value = (os.getenv(name) or "").strip()
    if value:
      return value
  raise RuntimeError(f"Missing required env var (one of): {', '.join(names)}")


def _retry(fn, *, tries: int = 3, base_sleep_s: float = 1.0):
  last_err: Optional[BaseException] = None
  for i in range(tries):
    try:
      return fn()
    except BaseException as e:  # noqa: BLE001
      last_err = e
      sleep_s = base_sleep_s * (2**i)
      time.sleep(sleep_s)
  assert last_err is not None
  raise last_err


def resolve_env_templates(value: str) -> str:
  # Railway sometimes provides templated variables like:
  # postgresql://...@${{RAILWAY_TCP_PROXY_DOMAIN}}:${{RAILWAY_TCP_PROXY_PORT}}/railway
  # These only resolve inside Railway. For local runs, we either substitute from env vars
  # (if provided) or raise a friendly error.
  if "${{" not in value:
    return value

  missing: List[str] = []

  def _repl(match: re.Match[str]) -> str:
    key = match.group(1)
    resolved = (os.getenv(key) or "").strip()
    if not resolved:
      missing.append(key)
      return match.group(0)
    return resolved

  rendered = ENV_TEMPLATE_RE.sub(_repl, value)
  if missing:
    raise RuntimeError(
      "DATABASE_URL contains Railway template placeholders that are not resolvable locally: "
      + ", ".join(sorted(set(missing)))
      + ".\nFix: set DATABASE_URL to Railway Postgres **Public connection string** (with real host/port), "
      + "or export those variables before running the script."
    )

  if "${{" in rendered:
    raise RuntimeError("DATABASE_URL still contains unresolved template placeholders after substitution.")

  return rendered


def sanitize_database_url_for_psycopg2(database_url: str) -> str:
  """
  psycopg2/libpq accepts many standard URI query params (e.g. sslmode),
  but NOT Prisma's `?schema=public` parameter.

  Railway / Prisma / Vercel often provide DATABASE_URL like:
    postgresql://.../railway?schema=public
  Strip only the unsupported `schema` param and keep everything else.
  """
  raw = (database_url or "").strip()
  if not raw:
    return raw
  if "?" not in raw:
    return raw

  parts = urlsplit(raw)
  query = parse_qsl(parts.query, keep_blank_values=True)
  filtered = [(k, v) for (k, v) in query if k.lower() != "schema"]
  if filtered == query:
    return raw

  rebuilt = urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(filtered), parts.fragment))
  return rebuilt


def _extract_json_object(text: str) -> Dict[str, Any]:
  try:
    return json.loads(text)
  except Exception:  # noqa: BLE001
    # Best-effort: trim code fences / surrounding commentary.
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
      return json.loads(text[start : end + 1])
    raise


def _get_first_candidate_text(payload: Dict[str, Any]) -> str:
  candidates = payload.get("candidates") or []
  if not candidates:
    raise RuntimeError(f"Gemini response missing candidates: {payload}")
  content = (candidates[0] or {}).get("content") or {}
  parts = content.get("parts") or []
  if not parts:
    raise RuntimeError(f"Gemini response missing content parts: {payload}")
  text = (parts[0] or {}).get("text")
  if not isinstance(text, str) or not text.strip():
    raise RuntimeError(f"Gemini response missing text: {payload}")
  return text


class GeminiClient:
  def __init__(self, *, api_key: str, api_base_url: str):
    self._api_key = api_key
    self._api_base_url = api_base_url.rstrip("/")
    self._session = requests.Session()
    self._session.headers.update({"Content-Type": "application/json"})

  @staticmethod
  def normalize_model_name(model: str) -> str:
    # ListModels returns names like "models/gemini-2.5-flash". Our URL already includes "/models/".
    # Accept either "gemini-2.5-flash" or "models/gemini-2.5-flash".
    model = model.strip()
    if model.startswith("models/"):
      return model[len("models/") :]
    return model

  def list_models(self) -> List[Dict[str, Any]]:
    url = f"{self._api_base_url}/models"

    def _call():
      resp = self._session.get(url, params={"key": self._api_key}, timeout=60)
      if resp.status_code >= 400:
        raise RuntimeError(f"Gemini ListModels failed ({resp.status_code}): {resp.text[:500]}")
      return resp.json()

    payload = _retry(_call, tries=3, base_sleep_s=1.0)
    models = payload.get("models") or []
    if not isinstance(models, list):
      raise RuntimeError(f"Gemini ListModels response invalid: {payload}")
    return models

  def generate_json(self, *, model: str, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
    model_id = self.normalize_model_name(model)
    url = f"{self._api_base_url}/models/{model_id}:generateContent"
    body = {
      "systemInstruction": {
        "parts": [
          {
            "text": system_prompt
            + "\n\nOutput MUST be valid JSON (no markdown, no code fences, double quotes only, no trailing commas)."
          }
        ]
      },
      "contents": [{"role": "user", "parts": [{"text": user_prompt}]}],
      "generationConfig": {
        "temperature": 0.0,
        "responseMimeType": "application/json",
        # Give the model enough room; malformed JSON often comes from truncation.
        "maxOutputTokens": 2048,
      },
    }

    def _call():
      resp = self._session.post(url, params={"key": self._api_key}, json=body, timeout=60)
      if resp.status_code >= 400:
        if resp.status_code == 404:
          raise RuntimeError(
            f"Gemini generateContent failed (404): model '{model}' not found / not supported on '{self._api_base_url}'. "
            f"Tip: run with --list-models to see available models, or set GEMINI_API_BASE_URL "
            f"to switch API version (e.g. https://generativelanguage.googleapis.com/v1). "
            f"Raw: {resp.text[:400]}"
          )
        raise RuntimeError(f"Gemini generateContent failed ({resp.status_code}): {resp.text[:500]}")
      payload = resp.json()
      text = _get_first_candidate_text(payload)
      # Parse inside retry so transient malformed outputs can be retried.
      return _extract_json_object(text)

    data = _retry(_call, tries=3, base_sleep_s=1.0)
    if not isinstance(data, dict):
      raise RuntimeError(f"Gemini JSON output is not an object: {type(data)}")
    return data

  def embed_text(self, *, model: str, text: str) -> List[float]:
    model_id = self.normalize_model_name(model)
    url = f"{self._api_base_url}/models/{model_id}:embedContent"
    body = {"content": {"parts": [{"text": text}]}}

    def _call():
      resp = self._session.post(url, params={"key": self._api_key}, json=body, timeout=60)
      if resp.status_code >= 400:
        raise RuntimeError(f"Gemini embedContent failed ({resp.status_code}): {resp.text[:500]}")
      return resp.json()

    payload = _retry(_call, tries=3, base_sleep_s=1.0)
    embedding = (payload.get("embedding") or {}).get("values")
    if not isinstance(embedding, list) or not embedding:
      raise RuntimeError(f"Gemini embedding missing values: {payload}")
    return [float(x) for x in embedding]


def normalize_embedding_dim(embedding: List[float], *, dim: int) -> List[float]:
  if len(embedding) == dim:
    return embedding
  if len(embedding) < dim:
    # Zero-padding keeps cosine similarity identical in the original subspace.
    return embedding + [0.0] * (dim - len(embedding))
  # Truncate to fit the DB column (vector(dim)). This is an MVP trade-off.
  # If you want full fidelity, migrate the DB column to vector(len(embedding)).
  print(f"⚠️ Embedding dim {len(embedding)} > {dim}; truncating to {dim}.")
  return embedding[:dim]


def _clamp_int(value: Any, *, min_value: int, max_value: int) -> int:
  try:
    n = int(float(value))
  except Exception:  # noqa: BLE001
    n = min_value
  if n < min_value:
    return min_value
  if n > max_value:
    return max_value
  return n


def _clamp_float(value: Any, *, min_value: float, max_value: float) -> float:
  try:
    n = float(value)
  except Exception:  # noqa: BLE001
    n = min_value
  if n < min_value:
    return min_value
  if n > max_value:
    return max_value
  return n


def _coerce_keywords(value: Any) -> List[str]:
  if isinstance(value, list):
    out = [str(x).strip() for x in value if str(x).strip()]
    return out[:8]
  if isinstance(value, str) and value.strip():
    # Allow comma-separated strings.
    parts = [p.strip() for p in value.split(",")]
    return [p for p in parts if p][:8]
  return []


def get_social_simulation_openai(
  *,
  brand: str,
  name: str,
  ingredients_text: Optional[str],
  api_key: str,
  model: str,
  api_base_url: str,
) -> Dict[str, Any]:
  url = api_base_url.rstrip("/") + "/chat/completions"

  system_prompt = SOCIAL_PROMPT.format(brand=brand, name=name)
  user_prompt = f"Product: {brand} - {name}\n"
  if ingredients_text:
    user_prompt += f"Ingredients: {ingredients_text}\n"

  body = {
    "model": model,
    "temperature": 0.2,
    "response_format": {"type": "json_object"},
    "messages": [
      {"role": "system", "content": system_prompt},
      {"role": "user", "content": user_prompt},
    ],
  }

  headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

  def _call():
    resp = requests.post(url, headers=headers, json=body, timeout=60)
    if resp.status_code >= 400:
      raise RuntimeError(f"OpenAI chat.completions failed ({resp.status_code}): {resp.text[:500]}")
    payload = resp.json()
    content = ((payload.get("choices") or [{}])[0].get("message") or {}).get("content")
    if not isinstance(content, str) or not content.strip():
      raise RuntimeError(f"OpenAI response missing content: {payload}")
    data = _extract_json_object(content)
    if not isinstance(data, dict):
      raise RuntimeError("OpenAI social simulation output is not an object")
    return data

  raw = _retry(_call, tries=3, base_sleep_s=1.0)

  # Normalize to internal schema expected by DB insert (snake_case), keep both keys for debugging if needed.
  return {
    "red_score": _clamp_int(raw.get("redScore"), min_value=0, max_value=100),
    "reddit_score": _clamp_int(raw.get("redditScore"), min_value=0, max_value=100),
    "burn_rate": _clamp_float(raw.get("burnRate"), min_value=0.0, max_value=1.0),
    "top_keywords": _coerce_keywords(raw.get("topKeywords")),
    "_raw": raw,
  }


def get_vectors_from_llm(client: GeminiClient, *, model: str, brand: str, name: str, ingredients: str) -> Dict[str, Any]:
  prompt = f"Product: {brand} {name}\nIngredients: {ingredients}"

  data = client.generate_json(model=model, system_prompt=SYSTEM_PROMPT, user_prompt=prompt)

  mechanism = data.get("mechanism") or {}
  risk_flags = data.get("risk_flags") or []
  experience = data.get("experience_prediction") or {}
  social = data.get("social_stats") or {}

  if not isinstance(mechanism, dict):
    raise ValueError("LLM output invalid: `mechanism` must be an object")
  if not isinstance(risk_flags, list):
    raise ValueError("LLM output invalid: `risk_flags` must be a list")
  if not isinstance(experience, dict):
    raise ValueError("LLM output invalid: `experience_prediction` must be an object")
  if social and not isinstance(social, dict):
    raise ValueError("LLM output invalid: `social_stats` must be an object when present")

  def _clamp_int_0_100(value: Any) -> int:
    try:
      n = int(round(float(value)))
    except Exception:  # noqa: BLE001
      return 0
    return max(0, min(100, n))

  def _clamp_float_0_1(value: Any) -> float:
    try:
      n = float(value)
    except Exception:  # noqa: BLE001
      return 0.0
    if n < 0:
      return 0.0
    if n > 1:
      return 1.0
    return n

  cleaned_risk_flags = [str(x).strip() for x in risk_flags if str(x).strip()]
  final_risk_flags = postprocess_risk_flags(llm_flags=cleaned_risk_flags, product_name=name, ingredients_text=ingredients)

  def _default_burn_rate(flags: List[str]) -> float:
    lower = [f.lower() for f in flags]
    if any(f in lower for f in ("high_irritation", "strong_acid", "retinol_high", "benzoyl_peroxide")):
      return 0.15
    if any(f in lower for f in ("alcohol_high", "fragrance", "mint", "mild_acid")):
      return 0.08
    return 0.03

  # Social stats are often not available from ingredients alone; we accept LLM-provided estimates
  # but also provide safe defaults so the pipeline is usable out-of-the-box.
  red_score = _clamp_int_0_100((social.get("red_score") or social.get("redScore") or 60) if isinstance(social, dict) else 60)
  reddit_score = _clamp_int_0_100((social.get("reddit_score") or social.get("redditScore") or 60) if isinstance(social, dict) else 60)
  burn_rate_raw = (
    (social.get("burn_rate") or social.get("burnRate") or _default_burn_rate(final_risk_flags)) if isinstance(social, dict) else _default_burn_rate(final_risk_flags)
  )
  burn_rate = calibrate_burn_rate(burn_rate=_clamp_float_0_1(burn_rate_raw), risk_flags=final_risk_flags, product_name=name)

  top_keywords_raw = social.get("top_keywords", []) if isinstance(social, dict) else []
  if isinstance(top_keywords_raw, str):
    top_keywords = [t.strip() for t in top_keywords_raw.split(",") if t.strip()]
  elif isinstance(top_keywords_raw, list):
    top_keywords = [str(t).strip() for t in top_keywords_raw if str(t).strip()]
  else:
    top_keywords = []

  normalized = {
    "mechanism": {
      "oil_control": _clamp_int_0_100(mechanism.get("oil_control", 0)),
      "anti_aging": _clamp_int_0_100(mechanism.get("anti_aging", 0)),
      "soothing": _clamp_int_0_100(mechanism.get("soothing", 0)),
      "barrier_repair": _clamp_int_0_100(mechanism.get("barrier_repair", 0)),
    },
    "risk_flags": final_risk_flags,
    "experience_prediction": {
      "texture": str(experience.get("texture", "")).strip() or "unknown",
      "finish": str(experience.get("finish", "")).strip() or "natural",
    },
    "social_stats": {
      "red_score": red_score,
      "reddit_score": reddit_score,
      "burn_rate": burn_rate,
      "top_keywords": top_keywords[:20],
    },
  }

  return normalized


def get_embedding(client: GeminiClient, *, model: str, text: str) -> List[float]:
  floats = client.embed_text(model=model, text=text)
  return normalize_embedding_dim(floats, dim=DEFAULT_EMBEDDING_DIM)


def embedding_to_vector_literal(embedding: List[float]) -> str:
  # pgvector accepts a text literal like: '[0.1, -0.2, ...]'::vector
  return "[" + ",".join(f"{x:.8f}" for x in embedding) + "]"


def parse_ingredients_list(text: str) -> List[str]:
  # Keep it simple: split by comma; Excel sheets commonly store INCI as comma-separated.
  items = [t.strip() for t in text.split(",")]
  return [t for t in items if t]


def _looks_like_wash_off_product(name: str) -> bool:
  n = (name or "").lower()
  return any(
    k in n
    for k in (
      "cleanser",
      "cleansing",
      "face wash",
      "facial wash",
      "wash",
      "soap",
      "foaming",
      "foam",
      "gel cleanser",
      "body wash",
    )
  )


def infer_risk_flags_from_ingredients(*, product_name: str, ingredients_text: str) -> List[str]:
  """
  Deterministic, conservative risk flag extraction.

  Goal:
  - Avoid LLM hallucination/over-flagging (e.g., mild cleansers marked as high_irritation).
  - Emit a controlled vocabulary used by downstream VETO logic.
  """
  text = unicodedata.normalize("NFKC", str(ingredients_text or "")).lower()
  items = [unicodedata.normalize("NFKC", t).lower() for t in parse_ingredients_list(text)]
  top5 = items[:5]
  top10 = items[:10]

  def _has_any(haystack: List[str], needles: List[str]) -> bool:
    return any(any(n in token for n in needles) for token in haystack)

  # Alcohol (denat) in top ingredients is a strong irritation signal for impaired barriers.
  alcohol_terms = ["alcohol denat", "alcohol denat.", "denatured alcohol", "sd alcohol", "ethyl alcohol"]
  has_alcohol_high = _has_any(top5, alcohol_terms)

  # Retinoids (leave-on "actives" category).
  retinoid_terms = [
    "retinol",
    "retinal",
    "retinaldehyde",
    "tretinoin",
    "adapalene",
    "tazarotene",
    "retinoate",
    "hydroxypinacolone retinoate",
  ]
  has_retinoid = _has_any(items, retinoid_terms)

  # Strong acids: treat BHA and strong AHA as "strong_acid". Exclude common pH adjusters like citric acid.
  strong_acid_terms = [
    "salicylic acid",
    "capryloyl salicylic acid",
    "betaine salicylate",
    "glycolic acid",
    "lactic acid",
  ]
  has_strong_acid = _has_any(top10, strong_acid_terms)

  # Mild acids: useful for comedones even in sensitive users; should NOT trigger "strong acid" veto.
  mild_acid_terms = ["azelaic acid", "mandelic acid", "gluconolactone", "lactobionic acid", "pha"]
  has_mild_acid = _has_any(items, mild_acid_terms)

  # Benzoyl peroxide is a high-irritation treatment class.
  has_benzoyl_peroxide = _has_any(items, ["benzoyl peroxide"])

  # Fungal acne heuristics (simple MVP): polysorbates.
  has_polysorbate = _has_any(items, ["polysorbate"])

  # Fragrance & common EU allergen markers (often indicate fragrance load).
  fragrance_markers = ["fragrance", "parfum", "perfume"]
  fragrance_allergens = [
    "limonene",
    "linalool",
    "citral",
    "geraniol",
    "eugenol",
    "coumarin",
    "farnesol",
    "benzyl benzoate",
    "benzyl salicylate",
  ]
  has_fragrance = _has_any(items, fragrance_markers) or _has_any(items, fragrance_allergens)

  # Mint/cooling agents (often sting on impaired barriers).
  has_mint = _has_any(items, ["menthol", "peppermint", "mentha", "camphor", "eucalyptus"])

  flags: List[str] = []
  if has_alcohol_high:
    flags.append("alcohol_high")
  if has_polysorbate:
    flags.append("fungal_acne")
  if has_fragrance:
    flags.append("fragrance")
  if has_mint:
    flags.append("mint")
  if has_retinoid:
    flags.append("retinol_high")
  if has_benzoyl_peroxide:
    flags.append("benzoyl_peroxide")
  if has_strong_acid:
    flags.append("strong_acid")
  if has_mild_acid and not has_strong_acid:
    # Keep distinct from strong acids so VETO can be nuanced.
    flags.append("mild_acid")

  # High-level "this can sting" summary flag. Only set when we have strong actives.
  if has_retinoid or has_strong_acid or has_benzoyl_peroxide:
    flags.append("high_irritation")

  # Avoid duplications while keeping a stable order for diffs/debugging.
  order = [
    "alcohol_high",
    "strong_acid",
    "mild_acid",
    "retinol_high",
    "benzoyl_peroxide",
    "high_irritation",
    "fragrance",
    "mint",
    "fungal_acne",
  ]
  stable = [f for f in order if f in set(flags)]
  for f in flags:
    if f not in stable:
      stable.append(f)
  return stable


def infer_key_actives_from_ingredients(*, ingredients_text: str, expert_key_actives: Optional[str] = None) -> List[str]:
  """
  Deterministic key-active extraction for KB/RAG grounding.

  Goal:
  - Provide a conservative, explainable "key_actives" list even when comparison sheets are missing.
  - Avoid brittle parsing; prefer substring hits for common, high-signal actives.
  """
  text = unicodedata.normalize("NFKC", str(ingredients_text or ""))
  lower = text.lower()

  out: List[str] = []
  seen = set()

  def _add(label: str) -> None:
    key = label.lower()
    if key in seen:
      return
    seen.add(key)
    out.append(label)

  def _has_any(needles: List[str]) -> bool:
    return any(n.lower() in lower for n in needles)

  # 1) Seed from any provided expert key_actives (already curated from sheets).
  if expert_key_actives:
    for part in re.split(r"[|,，;/]+", str(expert_key_actives)):
      cleaned = part.strip()
      if len(cleaned) < 2:
        continue
      if cleaned.lower() in {"n/a", "na", "none", "unknown"}:
        continue
      _add(cleaned)

  # 2) Ingredient-derived actives (small controlled list).
  if _has_any(["niacinamide", "烟酰胺"]):
    _add("Niacinamide")
  if _has_any(["tranexamic acid", "传明酸"]):
    _add("Tranexamic Acid")
  if _has_any(["alpha-arbutin", "arbutin", "熊果苷"]):
    _add("Arbutin")
  if _has_any(["kojic", "曲酸"]):
    _add("Kojic Acid")
  if _has_any(["azelaic", "壬二酸"]):
    _add("Azelaic Acid")

  # Vitamin C family: multiple derivatives; keep the label generic.
  if _has_any(["ascorbic acid", "l-ascorbic", "ascorbyl", "ascorbate", "维c", "维生素c"]):
    _add("Vitamin C (Ascorbate family)")

  # Retinoids
  if _has_any(["retinol", "retinal", "retinaldehyde", "tretinoin", "adapalene", "retinoate", "a醇", "维a", "视黄"]):
    _add("Retinoid")

  # Acids
  if _has_any(["salicylic acid", "betaine salicylate", "capryloyl salicylic", "水杨酸"]):
    _add("BHA (Salicylic Acid)")
  if _has_any(["glycolic acid", "lactic acid", "乙醇酸", "乳酸"]):
    _add("AHA (Glycolic/Lactic)")
  if _has_any(["mandelic acid", "杏仁酸"]):
    _add("Mandelic Acid")
  if _has_any(["gluconolactone", "lactobionic", "pha", "葡糖酸内酯"]):
    _add("PHA")

  # Peptides (broad)
  if _has_any(["peptide", "tripeptide", "hexapeptide", "palmitoyl", "ghk", "copper tripeptide", "多肽", "蓝铜"]):
    _add("Peptides")

  # Barrier / soothing staples
  if _has_any(["panthenol", "d-panthenol", "泛醇"]):
    _add("Panthenol (B5)")
  if _has_any(["ceramide", "神经酰胺"]):
    _add("Ceramides")
  if _has_any(["cholesterol", "胆固醇"]):
    _add("Cholesterol")
  if _has_any(["centella", "madecassoside", "asiaticoside", "积雪草"]):
    _add("Centella (Madecassoside family)")
  if _has_any(["allantoin", "尿囊素"]):
    _add("Allantoin")
  if _has_any(["hyaluronic", "sodium hyaluronate", "透明质酸", "玻尿酸"]):
    _add("Hyaluronic Acid")

  # Acne treatments
  if _has_any(["benzoyl peroxide", "过氧化苯甲酰"]):
    _add("Benzoyl Peroxide")
  if _has_any(["adapalene", "阿达帕林"]):
    _add("Adapalene")

  # Keep it compact for RAG/token budgets.
  return out[:16]


def postprocess_risk_flags(
  *, llm_flags: List[str], product_name: str, ingredients_text: str
) -> List[str]:
  """
  Use LLM risk flags as hints, but enforce deterministic evidence + controlled vocabulary.

  This prevents false positives that cascade into safety VETO (e.g., "high_irritation" on a basic cleanser).
  """
  deterministic = infer_risk_flags_from_ingredients(product_name=product_name, ingredients_text=ingredients_text)
  det_set = set(deterministic)

  def _norm_flag(flag: str) -> str:
    s = unicodedata.normalize("NFKC", str(flag or "")).strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "", s)
    return s

  # Only keep LLM flags that are (a) in our allowlist and (b) supported by deterministic signals.
  allowlist = {
    "alcohol_high",
    "strong_acid",
    "mild_acid",
    "retinol_high",
    "benzoyl_peroxide",
    "high_irritation",
    "fungal_acne",
    "fragrance",
    "mint",
  }

  # Many models output generic flags like "alcohol"/"acid"/"retinol" which are too broad.
  # Map them to our stricter tokens only if deterministic evidence exists; otherwise drop.
  synonym_map = {
    "alcohol": "alcohol_high",
    "high_alcohol": "alcohol_high",
    "acid": "strong_acid",
    "aha": "strong_acid",
    "bha": "strong_acid",
    "retinol": "retinol_high",
    "retinoid": "retinol_high",
  }

  cleaned: List[str] = []
  for raw in llm_flags:
    f = _norm_flag(raw)
    if not f:
      continue
    f = synonym_map.get(f, f)
    if f not in allowlist:
      continue
    # High-risk flags must be backed by deterministic evidence.
    if f in ("high_irritation", "strong_acid", "mild_acid", "retinol_high", "benzoyl_peroxide", "alcohol_high", "fungal_acne", "fragrance", "mint"):
      if f not in det_set:
        continue
    cleaned.append(f)

  merged = list(dict.fromkeys(deterministic + cleaned))
  return merged


def calibrate_burn_rate(*, burn_rate: float, risk_flags: List[str], product_name: str) -> float:
  """
  Calibrate burn_rate so it is consistent with deterministic risk flags.

  Why: burn_rate > 0.10 triggers strong VETO in chat prompts/engines. For wash-off cleansers without strong actives,
  cap burn_rate to avoid over-vetoing the entire catalog.
  """
  br = _clamp_float(burn_rate, min_value=0.0, max_value=1.0)
  wash_off = _looks_like_wash_off_product(product_name)

  has_strong = any(f in risk_flags for f in ("strong_acid", "retinol_high", "benzoyl_peroxide", "high_irritation"))
  has_medium = any(f in risk_flags for f in ("alcohol_high", "fragrance", "mint"))

  if wash_off and not has_strong:
    # Keep cleansers usable for sensitive users unless we have clear strong-actives evidence.
    cap = 0.10 if has_medium else 0.08
    return min(br, cap)

  if has_strong:
    # Allow higher irritation reports for strong leave-on actives.
    return max(br, 0.12)

  if has_medium:
    return min(br, 0.25)

  return min(br, 0.15)


def split_brand_and_name(product_full_name: str) -> Tuple[str, str]:
  """
  Best-effort brand/name split for sheets that only have a single "Product" column.

  Strategy:
  - Prefer matching known multi-word brands (longest-first).
  - Fallback to first token as brand.
  - Keep name as the remainder (or full string if unknown).
  """
  full = (product_full_name or "").strip()
  if not full:
    return ("Unknown", "")

  known_multi = sorted(
    [
      "La Roche-Posay",
      "Paula's Choice",
      "The Ordinary",
      "Estée Lauder",
      "Estee Lauder",
      "First Aid Beauty",
      "Beauty of Joseon",
      "Helena Rubinstein",
      "SkinCeuticals",
      "La Mer",
      "Tom Ford",
      "Hada Labo",
    ],
    key=len,
    reverse=True,
  )

  lower = full.lower()
  for b in known_multi:
    if lower.startswith(b.lower()):
      rest = full[len(b) :].strip()
      rest = rest.lstrip("-–—:").strip()
      return (b, rest or full)

  # Fallback: first token as brand.
  first = full.split(" ")[0].strip()
  if not first:
    return ("Unknown", full)
  rest = full[len(first) :].strip()
  rest = rest.lstrip("-–—:").strip()
  return (first, rest or full)


def safe_split_brand_and_name(product_full_name: str) -> Tuple[str, str]:
  """
  Split brand/name, but fall back to the full string as name if split fails.
  """
  b, n = split_brand_and_name(product_full_name)
  b = (b or "").strip() or "Unknown"
  n = (n or "").strip() or (product_full_name or "").strip()
  return b, n


def canonicalize_kb_field(label: str) -> str:
  """
  Produce a stable, unique-ish column key for KB snippets.

  Notes:
  - We keep this key per sheet/column (used in UNIQUE constraint), so it MUST
    stay stable across runs.
  - Some sheets contain Chinese-only column names; the previous ASCII-only
    normalization collapsed these into "unknown" and caused collisions.
  """
  raw = (label or "").strip()
  if not raw:
    return "unknown"

  value = raw.lower()
  value = re.sub(r"[^a-z0-9]+", "_", value)
  value = value.strip("_")
  if value:
    return value[:64]

  # Fallback for non-ASCII labels: use a short stable hash.
  digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
  return f"col_{digest}"


def infer_kb_canonical_key(label: str) -> Optional[str]:
  """
  Map heterogeneous spreadsheet column labels to a small, stable ontology.

  Stored in KB snippet `metadata.canonical_key` (no DB schema change required).
  """
  raw = (label or "").strip()
  if not raw:
    return None

  lower = raw.lower()

  # Sensitivity / irritation
  if (
    "sensitivity" in lower
    or "irrit" in lower
    or "risk" in lower
    or "敏感" in raw
    or "刺激" in raw
    or "刺痛" in raw
    or "过敏" in raw
  ):
    return "sensitivity"

  # Key actives
  if (
    "key_actives" in lower
    or ("key" in lower and ("active" in lower or "actives" in lower))
    or "核心成分" in raw
    or "主要成分" in raw
    or "关键活性" in raw
    or "功效成分" in raw
    or "活性" in raw
  ):
    return "key_actives"

  # Comparison / dupes
  if (
    "comparison" in lower
    or "compare" in lower
    or "dupe" in lower
    or "替代" in raw
    or "平替" in raw
    or "对比" in raw
    or "竞品" in raw
  ):
    return "comparison"

  # Usage / pairing
  if (
    "usage" in lower
    or "routine" in lower
    or "layer" in lower
    or "frequency" in lower
    or "用法" in raw
    or "搭配" in raw
    or "叠加" in raw
    or "频率" in raw
    or "注意事项" in raw
  ):
    return "usage"

  # Texture / finish
  if (
    "texture" in lower
    or "finish" in lower
    or "pilling" in lower
    or "质地" in raw
    or "清爽" in raw
    or "厚重" in raw
    or "搓泥" in raw
    or "成膜" in raw
    or "油腻" in raw
  ):
    return "texture"

  if "notes" in lower or "note" in lower or "备注" in raw or "评价" in raw:
    return "notes"

  return None


def normalize_match_key(text: str) -> str:
  value = unicodedata.normalize("NFKD", str(text or ""))
  value = "".join(ch for ch in value if not unicodedata.combining(ch))
  value = value.lower()
  value = re.sub(r"[^a-z0-9]+", "", value)
  return value


def _cell_text(value: Any) -> str:
  if value is None:
    return ""
  s = str(value).strip()
  if not s or s.lower() == "nan":
    return ""
  return s


def extract_kb_snippets_from_workbook(*, path: str) -> List[KbSnippet]:
  """
  Extract non-ingredient notes from all sheets in the workbook.

  We intentionally skip Ingredients/Source columns (already stored elsewhere).
  """
  wb = load_workbook(path, read_only=True, data_only=True)
  out: List[KbSnippet] = []

  for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
      continue

    headers = [str(h).strip() if h is not None else "" for h in header]
    headers_lc = [h.strip().lower() for h in headers]

    if "product" not in headers_lc:
      continue
    idx_product = headers_lc.index("product")
    idx_source = headers_lc.index("source") if "source" in headers_lc else None

    skip = {"product", "ingredients (as listed)", "ingredients", "source"}
    note_cols: List[Tuple[int, str]] = []
    for j, label in enumerate(headers):
      if not label.strip():
        continue
      if label.strip().lower() in skip:
        continue
      note_cols.append((j, label))

    if not note_cols:
      continue

    for row_number, r in enumerate(rows, start=2):
      if not r or idx_product >= len(r):
        continue
      product_full = _cell_text(r[idx_product])
      if not product_full:
        continue
      brand, name = split_brand_and_name(product_full)

      source_url = ""
      if idx_source is not None and idx_source < len(r):
        source_url = _cell_text(r[idx_source])

      for j, label in note_cols:
        if j >= len(r):
          continue
        content = _cell_text(r[j])
        if len(content) < 2:
          continue
        canonical_key = infer_kb_canonical_key(label)
        meta: Dict[str, Any] = {
          "source_file": os.path.basename(path),
          "source_sheet": sheet_name,
          "field_label": label,
          "row_number": row_number,
          "product_full_name": product_full,
        }
        if canonical_key:
          meta["canonical_key"] = canonical_key
          meta["canonical_key_source"] = "label_rules_v1"
        if source_url:
          meta["source"] = source_url

        out.append(
          KbSnippet(
            brand=brand,
            name=name,
            source_sheet=sheet_name,
            field=canonicalize_kb_field(label),
            content=content,
            metadata=meta,
          )
        )

  return out


class AuroraDb:
  def __init__(self, database_url: str):
    self._database_url = database_url
    self._has_region_availability = False
    self._has_kb_snippets_table = False
    self._has_product_aliases_table = False
    self._product_ids_norm_cache: Optional[Dict[Tuple[str, str], str]] = None

  def __enter__(self):
    self.conn = psycopg2.connect(self._database_url)
    self.conn.autocommit = False
    self._has_region_availability = self._ensure_region_availability_column()
    self._has_kb_snippets_table = self._ensure_kb_snippets_table()
    self._has_product_aliases_table = self._ensure_product_aliases_table()
    self._product_ids_norm_cache = None
    return self

  def __exit__(self, exc_type, exc, tb):
    try:
      if exc:
        self.conn.rollback()
    finally:
      self.conn.close()

  def _ensure_region_availability_column(self) -> bool:
    """
    Best-effort schema guard.

    The ingestion JSON may include `availability: ["CN","US"]`. We store it in
    products.region_availability (TEXT[]). If the column does not exist yet,
    attempt to add it. If we cannot, we degrade gracefully and skip writing it.
    """
    try:
      with self.conn.cursor() as cur:
        cur.execute(
          """
          SELECT 1
          FROM information_schema.columns
          WHERE table_schema = 'public'
            AND table_name = 'products'
            AND column_name = 'region_availability'
          LIMIT 1;
          """
        )
        if cur.fetchone():
          return True

        cur.execute(
          'ALTER TABLE "products" ADD COLUMN IF NOT EXISTS region_availability TEXT[] NOT NULL DEFAULT \'{}\'::text[];'
        )
        self.conn.commit()
        return True
    except BaseException:  # noqa: BLE001
      try:
        self.conn.rollback()
      except BaseException:  # noqa: BLE001
        pass
      print("⚠️  Could not ensure products.region_availability column; continuing without it.")
      return False

  def _ensure_kb_snippets_table(self) -> bool:
    try:
      with self.conn.cursor() as cur:
        cur.execute(
          """
          CREATE TABLE IF NOT EXISTS "product_kb_snippets" (
            id UUID PRIMARY KEY,
            product_id UUID NOT NULL REFERENCES "products"(id) ON DELETE CASCADE,
            source_sheet TEXT NOT NULL,
            field TEXT NOT NULL,
            content TEXT NOT NULL,
            metadata JSONB,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            CONSTRAINT product_kb_snippets_product_sheet_field_uidx UNIQUE (product_id, source_sheet, field)
          );
          """
        )
        # If the table existed without the unique constraint, ensure a unique index exists so
        # our ON CONFLICT (product_id, source_sheet, field) upsert works.
        cur.execute(
          'CREATE UNIQUE INDEX IF NOT EXISTS product_kb_snippets_product_sheet_field_uidx ON "product_kb_snippets"(product_id, source_sheet, field);'
        )
        cur.execute('CREATE INDEX IF NOT EXISTS product_kb_snippets_product_id_idx ON "product_kb_snippets"(product_id);')
      self.conn.commit()
      return True
    except BaseException:  # noqa: BLE001
      try:
        self.conn.rollback()
      except BaseException:  # noqa: BLE001
        pass
      print("⚠️  Could not ensure product_kb_snippets table; continuing without KB ingestion.")
      return False

  def _ensure_product_aliases_table(self) -> bool:
    """
    Best-effort table for DB-managed product aliases (anchor resolution).

    Production should use Prisma migrations; this is a safety net for local runs.
    """
    try:
      with self.conn.cursor() as cur:
        cur.execute(
          """
          CREATE TABLE IF NOT EXISTS "product_aliases" (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL REFERENCES "products"(id) ON DELETE CASCADE,
            alias TEXT NOT NULL,
            alias_normalized TEXT NOT NULL,
            kind TEXT,
            weight INTEGER NOT NULL DEFAULT 0,
            locale TEXT,
            created_at TIMESTAMP(3) NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP(3) NOT NULL
          );
          """
        )
        cur.execute(
          'CREATE UNIQUE INDEX IF NOT EXISTS product_aliases_product_alias_norm_uidx ON "product_aliases"(product_id, alias_normalized);'
        )
        cur.execute('CREATE INDEX IF NOT EXISTS product_aliases_alias_normalized_idx ON "product_aliases"(alias_normalized);')
        cur.execute('CREATE INDEX IF NOT EXISTS product_aliases_product_id_idx ON "product_aliases"(product_id);')
      self.conn.commit()
      return True
    except BaseException:  # noqa: BLE001
      try:
        self.conn.rollback()
      except BaseException:  # noqa: BLE001
        pass
      print("⚠️  Could not ensure product_aliases table; continuing without alias upserts.")
      return False

  def find_product_id(self, *, brand: str, name: str) -> Optional[str]:
    with self.conn.cursor() as cur:
      cur.execute('SELECT id FROM "products" WHERE brand = %s AND name = %s LIMIT 1;', (brand, name))
      row = cur.fetchone()
      return row[0] if row else None

  def find_product_id_loose(self, *, brand: str, name: str) -> Optional[str]:
    """
    Match an existing product id using exact brand/name first, then a normalized fallback.

    This helps avoid duplicates when sources vary in punctuation/diacritics
    (e.g., "Lancôme" vs "Lancome") while remaining relatively strict.
    """
    pid = self.find_product_id(brand=brand, name=name)
    if pid:
      return pid

    if self._product_ids_norm_cache is None:
      self._product_ids_norm_cache = self.load_all_product_ids_normalized()
    key = (normalize_match_key(brand), normalize_match_key(name))
    return self._product_ids_norm_cache.get(key)

  def ensure_product_stub(self, *, brand: str, name: str, availability: Optional[List[str]] = None) -> str:
    """
    Ensure a Product row exists for (brand,name) even if vectors/ingredients are missing.

    This is used to attach KB snippets for products where INCI is not provided yet.
    """
    existing = self.find_product_id_loose(brand=brand, name=name)
    if existing:
      return str(existing)

    product_id = str(uuid.uuid4())
    sku = InputSku(
      brand=str(brand).strip() or "Unknown",
      name=str(name).strip(),
      ingredients_text="",
      price_usd=0.0,
      price_cny=0.0,
      availability=availability or ["Global"],
    )
    self.insert_product(sku, product_id=product_id)
    # Keep alias table warm for better anchor resolution in chat (best-effort).
    self.upsert_default_aliases_for_product(product_id=product_id, brand=sku.brand, name=sku.name)
    # Refresh cache so subsequent lookups find it.
    self._product_ids_norm_cache = None
    return product_id

  def load_all_product_ids(self) -> Dict[Tuple[str, str], str]:
    with self.conn.cursor() as cur:
      cur.execute('SELECT brand, name, id FROM "products";')
      rows = cur.fetchall()
      out: Dict[Tuple[str, str], str] = {}
      for brand, name, pid in rows:
        out[(str(brand), str(name))] = str(pid)
      return out

  def load_all_products_basic(self) -> List[Tuple[str, str, str]]:
    with self.conn.cursor() as cur:
      cur.execute('SELECT id, brand, name FROM "products" ORDER BY updated_at DESC;')
      rows = cur.fetchall()
      return [(str(pid), str(brand), str(name)) for (pid, brand, name) in rows]

  def load_all_product_ids_normalized(self) -> Dict[Tuple[str, str], str]:
    with self.conn.cursor() as cur:
      cur.execute('SELECT brand, name, id FROM "products";')
      rows = cur.fetchall()
      out: Dict[Tuple[str, str], str] = {}
      for brand, name, pid in rows:
        key = (normalize_match_key(str(brand)), normalize_match_key(str(name)))
        # If collisions occur, keep the first seen to avoid flapping.
        out.setdefault(key, str(pid))
      return out

  def delete_product(self, product_id: str) -> None:
    with self.conn.cursor() as cur:
      cur.execute('DELETE FROM "products" WHERE id = %s;', (product_id,))

  def upsert_product(self, sku: InputSku, *, product_id: str) -> None:
    with self.conn.cursor() as cur:
      if self._has_region_availability:
        availability = list(sku.availability or [])
        cur.execute(
          """
          INSERT INTO "products" (
            id,
            brand,
            name,
            price_usd,
            price_cny,
            product_url,
            image_url,
            region_availability,
            created_at,
            updated_at
          )
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
          ON CONFLICT (id) DO UPDATE SET
            brand = EXCLUDED.brand,
            name = EXCLUDED.name,
            price_usd = EXCLUDED.price_usd,
            price_cny = EXCLUDED.price_cny,
            product_url = EXCLUDED.product_url,
            image_url = EXCLUDED.image_url,
            region_availability = EXCLUDED.region_availability,
            updated_at = NOW();
          """,
          (product_id, sku.brand, sku.name, sku.price_usd, sku.price_cny, sku.product_url, sku.image_url, availability),
        )
        return

      # Backward-compatible path if the DB column is missing.
      cur.execute(
        """
        INSERT INTO "products" (id, brand, name, price_usd, price_cny, product_url, image_url, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        ON CONFLICT (id) DO UPDATE SET
          brand = EXCLUDED.brand,
          name = EXCLUDED.name,
          price_usd = EXCLUDED.price_usd,
          price_cny = EXCLUDED.price_cny,
          product_url = EXCLUDED.product_url,
          image_url = EXCLUDED.image_url,
          updated_at = NOW();
        """,
        (product_id, sku.brand, sku.name, sku.price_usd, sku.price_cny, sku.product_url, sku.image_url),
      )

  def delete_vectors_for_product(self, product_id: str) -> None:
    with self.conn.cursor() as cur:
      cur.execute('DELETE FROM "sku_vectors" WHERE product_id = %s;', (product_id,))

  def delete_ingredients_for_product(self, product_id: str) -> None:
    with self.conn.cursor() as cur:
      cur.execute('DELETE FROM "ingredients" WHERE product_id = %s;', (product_id,))

  def delete_social_stats_for_product(self, product_id: str) -> None:
    with self.conn.cursor() as cur:
      cur.execute('DELETE FROM "social_stats" WHERE product_id = %s;', (product_id,))

  def insert_product(self, sku: InputSku, *, product_id: str) -> None:
    with self.conn.cursor() as cur:
      if self._has_region_availability:
        availability = list(sku.availability or [])
        cur.execute(
          """
          INSERT INTO "products" (
            id,
            brand,
            name,
            price_usd,
            price_cny,
            product_url,
            image_url,
            region_availability,
            created_at,
            updated_at
          )
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW());
          """,
          (product_id, sku.brand, sku.name, sku.price_usd, sku.price_cny, sku.product_url, sku.image_url, availability),
        )
        return

      # Backward-compatible path if the DB column is missing.
      cur.execute(
        """
        INSERT INTO "products" (id, brand, name, price_usd, price_cny, product_url, image_url, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW());
        """,
        (product_id, sku.brand, sku.name, sku.price_usd, sku.price_cny, sku.product_url, sku.image_url),
      )

  def insert_vectors(
    self,
    *,
    product_id: str,
    vector_id: str,
    mechanism: Dict[str, Any],
    experience: Dict[str, Any],
    risk_flags: List[str],
    embedding: Optional[List[float]],
  ) -> None:
    with self.conn.cursor() as cur:
      if embedding is None:
        cur.execute(
          """
          INSERT INTO "sku_vectors" (id, product_id, mechanism, experience, risk_flags, embedding)
          VALUES (%s, %s, %s, %s, %s, NULL);
          """,
          (vector_id, product_id, Json(mechanism), Json(experience), risk_flags),
        )
        return

      cur.execute(
        """
        INSERT INTO "sku_vectors" (id, product_id, mechanism, experience, risk_flags, embedding)
        VALUES (%s, %s, %s, %s, %s, %s::vector);
        """,
        (vector_id, product_id, Json(mechanism), Json(experience), risk_flags, embedding_to_vector_literal(embedding)),
      )

  def insert_ingredients(self, *, product_id: str, ingredient_id: str, full_list: List[str]) -> None:
    with self.conn.cursor() as cur:
      cur.execute(
        """
        INSERT INTO "ingredients" (id, product_id, full_list, hero_actives)
        VALUES (%s, %s, %s, %s);
        """,
        (ingredient_id, product_id, full_list, Json([])),
      )

  def insert_social_stats(
    self,
    *,
    product_id: str,
    social_id: str,
    red_score: int,
    reddit_score: int,
    burn_rate: float,
    top_keywords: List[str],
  ) -> None:
    with self.conn.cursor() as cur:
      cur.execute(
        """
        INSERT INTO "social_stats" (id, product_id, red_score, reddit_score, burn_rate, top_keywords, last_updated)
        VALUES (%s, %s, %s, %s, %s, %s, NOW());
        """,
        (social_id, product_id, red_score, reddit_score, burn_rate, top_keywords),
      )

  def upsert_kb_snippet(
    self,
    *,
    product_id: str,
    source_sheet: str,
    field: str,
    content: str,
    metadata: Dict[str, Any],
  ) -> None:
    if not self._has_kb_snippets_table:
      return
    with self.conn.cursor() as cur:
      cur.execute(
        """
        INSERT INTO "product_kb_snippets" (id, product_id, source_sheet, field, content, metadata, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
        ON CONFLICT (product_id, source_sheet, field) DO UPDATE SET
          content = EXCLUDED.content,
          metadata = EXCLUDED.metadata,
          updated_at = NOW();
        """,
        (str(uuid.uuid4()), product_id, source_sheet, field, content, Json(metadata)),
      )

  def upsert_product_alias(
    self,
    *,
    product_id: str,
    alias: str,
    kind: Optional[str],
    weight: int,
    locale: Optional[str] = None,
  ) -> None:
    if not self._has_product_aliases_table:
      return

    raw = str(alias or "").strip()
    if len(raw) < 2:
      return
    alias_norm = normalize_alias_text(raw)
    if len(alias_norm) < 2:
      return

    with self.conn.cursor() as cur:
      cur.execute(
        """
        INSERT INTO "product_aliases" (
          id, product_id, alias, alias_normalized, kind, weight, locale, created_at, updated_at
        ) VALUES (
          %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
        )
        ON CONFLICT (product_id, alias_normalized) DO UPDATE SET
          alias = EXCLUDED.alias,
          kind = EXCLUDED.kind,
          weight = EXCLUDED.weight,
          locale = EXCLUDED.locale,
          updated_at = NOW();
        """,
        (str(uuid.uuid4()), product_id, raw, alias_norm, kind, int(weight), locale),
      )

  def upsert_default_aliases_for_product(self, *, product_id: str, brand: str, name: str) -> None:
    if not self._has_product_aliases_table:
      return

    b = str(brand or "").strip()
    n = str(name or "").strip()
    if not b or not n:
      return

    full = f"{b} {n}".strip()

    aliases: List[Tuple[str, str, int]] = [
      (b, "brand", 10),
      (n, "name", 5),
      (full, "full_name", 20),
    ]

    brand_key = normalize_alias_text(b)
    for a in BRAND_ALIAS_SYNONYMS.get(brand_key, []):
      aliases.append((a, "brand_alias", 8))

    full_key = normalize_alias_text(full)
    for a in PRODUCT_NICKNAMES.get(full_key, []):
      aliases.append((a, "nickname", 50))

    for alias, kind, weight in aliases:
      self.upsert_product_alias(product_id=product_id, alias=alias, kind=kind, weight=weight, locale=None)


def _get_excel_rows(path: str, sheet: Optional[str]) -> Tuple[List[str], Iterable[List[Any]]]:
  wb = load_workbook(path, read_only=True, data_only=True)
  ws = wb[sheet] if sheet else wb.active

  rows = ws.iter_rows(values_only=True)
  header = next(rows, None)
  if not header:
    raise RuntimeError("Excel appears empty (missing header row)")

  headers = [str(h).strip() if h is not None else "" for h in header]
  return headers, rows


def _pick_column(headers: List[str], preferred: Optional[str], *, required: bool = True) -> Optional[int]:
  if preferred is None:
    return None
  lowered = [h.strip().lower() for h in headers]
  key = preferred.strip().lower()
  if key in lowered:
    return lowered.index(key)
  if required:
    raise RuntimeError(f"Missing required column: {preferred} (available: {headers})")
  return None


def load_input_skus_from_excel(
  *,
  path: str,
  sheet: Optional[str],
  col_brand: str,
  col_name: str,
  col_ingredients: str,
  col_price_usd: Optional[str],
  col_price: Optional[str],
  price_cny_rate: float,
  col_product_url: Optional[str],
  col_image_url: Optional[str],
  limit: Optional[int],
) -> List[InputSku]:
  headers, rows = _get_excel_rows(path, sheet)

  idx_brand = _pick_column(headers, col_brand)
  idx_name = _pick_column(headers, col_name)
  idx_ing = _pick_column(headers, col_ingredients)

  idx_price_usd = _pick_column(headers, col_price_usd, required=False)
  idx_price = _pick_column(headers, col_price, required=False)

  idx_product_url = _pick_column(headers, col_product_url, required=False)
  idx_image_url = _pick_column(headers, col_image_url, required=False)

  skus: List[InputSku] = []
  for r in rows:
    if r is None:
      continue
    brand = (r[idx_brand] if idx_brand < len(r) else "") or ""
    name = (r[idx_name] if idx_name < len(r) else "") or ""
    ingredients = (r[idx_ing] if idx_ing < len(r) else "") or ""
    if not str(brand).strip() or not str(name).strip() or not str(ingredients).strip():
      continue

    # Some source sheets only have a single "Product" column. If the user maps
    # both --col-brand and --col-name to that column, split brand/name here.
    if idx_brand == idx_name:
      inferred_brand, inferred_name = split_brand_and_name(str(brand).strip())
      brand = inferred_brand
      name = inferred_name

    price_usd: Optional[float] = None
    if idx_price_usd is not None:
      raw = r[idx_price_usd] if idx_price_usd < len(r) else None
      if raw is not None and str(raw).strip():
        price_usd = float(raw)

    if price_usd is None and idx_price is not None:
      raw = r[idx_price] if idx_price < len(r) else None
      if raw is not None and str(raw).strip():
        price_usd = float(raw)

    if price_usd is None:
      price_usd = 0.0

    product_url = None
    if idx_product_url is not None:
      raw = r[idx_product_url] if idx_product_url < len(r) else None
      if raw is not None and str(raw).strip():
        product_url = str(raw).strip()

    image_url = None
    if idx_image_url is not None:
      raw = r[idx_image_url] if idx_image_url < len(r) else None
      if raw is not None and str(raw).strip():
        image_url = str(raw).strip()

    skus.append(
      InputSku(
        brand=str(brand).strip(),
        name=str(name).strip(),
        ingredients_text=str(ingredients).strip(),
        price_usd=float(price_usd),
        price_cny=float(price_usd) * price_cny_rate,
        availability=["Global"],
        product_url=product_url,
        image_url=image_url,
      )
    )

    if limit is not None and len(skus) >= limit:
      break

  return skus


def load_input_skus_from_workbook_all_sheets(
  *,
  path: str,
  col_brand: str,
  col_name: str,
  col_ingredients: str,
  col_price_usd: Optional[str],
  col_price: Optional[str],
  price_cny_rate: float,
  col_product_url: Optional[str],
  col_image_url: Optional[str],
  limit: Optional[int],
) -> List[InputSku]:
  wb = load_workbook(path, read_only=True, data_only=True)

  out: List[InputSku] = []
  seen: set[Tuple[str, str]] = set()

  for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
      continue
    headers = [str(h).strip() if h is not None else "" for h in header]

    try:
      idx_brand = _pick_column(headers, col_brand)
      idx_name = _pick_column(headers, col_name)
      idx_ing = _pick_column(headers, col_ingredients)
      idx_price_usd = _pick_column(headers, col_price_usd, required=False)
      idx_price = _pick_column(headers, col_price, required=False)
      idx_product_url = _pick_column(headers, col_product_url, required=False)
      idx_image_url = _pick_column(headers, col_image_url, required=False)
    except BaseException as e:  # noqa: BLE001
      print(f"   ⚠️ Skipping sheet '{sheet_name}': {e}")
      continue

    for r in rows:
      if r is None:
        continue
      brand = (r[idx_brand] if idx_brand < len(r) else "") or ""
      name = (r[idx_name] if idx_name < len(r) else "") or ""
      ingredients = (r[idx_ing] if idx_ing < len(r) else "") or ""
      if not str(brand).strip() or not str(name).strip() or not str(ingredients).strip():
        continue

      if idx_brand == idx_name:
        inferred_brand, inferred_name = split_brand_and_name(str(brand).strip())
        brand = inferred_brand
        name = inferred_name

      brand_s = str(brand).strip()
      name_s = str(name).strip()
      key = (normalize_match_key(brand_s), normalize_match_key(name_s))
      if key in seen:
        continue
      seen.add(key)

      price_usd: Optional[float] = None
      if idx_price_usd is not None:
        raw = r[idx_price_usd] if idx_price_usd < len(r) else None
        if raw is not None and str(raw).strip():
          price_usd = float(raw)
      if price_usd is None and idx_price is not None:
        raw = r[idx_price] if idx_price < len(r) else None
        if raw is not None and str(raw).strip():
          price_usd = float(raw)
      if price_usd is None:
        price_usd = 0.0

      product_url = None
      if idx_product_url is not None:
        raw = r[idx_product_url] if idx_product_url < len(r) else None
        if raw is not None and str(raw).strip():
          product_url = str(raw).strip()

      image_url = None
      if idx_image_url is not None:
        raw = r[idx_image_url] if idx_image_url < len(r) else None
        if raw is not None and str(raw).strip():
          image_url = str(raw).strip()

      out.append(
        InputSku(
          brand=brand_s,
          name=name_s,
          ingredients_text=str(ingredients).strip(),
          price_usd=float(price_usd),
          price_cny=float(price_usd) * price_cny_rate,
          availability=["Global"],
          product_url=product_url,
          image_url=image_url,
        )
      )

      if limit is not None and len(out) >= limit:
        return out

  return out


def load_input_skus_from_json(*, path: str, price_cny_rate: float, limit: Optional[int]) -> List[InputSku]:
  with open(path, "r", encoding="utf-8") as f:
    payload = json.load(f)

  items = payload.get("items") if isinstance(payload, dict) else payload
  if not isinstance(items, list):
    raise RuntimeError("JSON input must be a list of items (or {\"items\": [...]}).")

  skus: List[InputSku] = []
  for i, raw in enumerate(items):
    if not isinstance(raw, dict):
      print(f"Skipping JSON row {i}: not an object")
      continue

    brand = str(raw.get("brand") or "").strip()
    name = str(raw.get("name") or "").strip()
    # Many CSV pipelines put the full product name (including brand) into `name`.
    # Attempt to split it to reduce duplicates when ingesting.
    if name and (not brand or name.lower().startswith(brand.lower())):
      split_brand, split_name = split_brand_and_name(name)
      if not brand and split_brand and split_brand != "Unknown":
        brand = split_brand
      if split_name:
        name = split_name
    ingredients = str(raw.get("ingredients_text") or raw.get("ingredients") or "").strip()

    expert_raw = raw.get("expert_knowledge") if isinstance(raw.get("expert_knowledge"), dict) else None
    kb_snippets_raw = raw.get("kb_snippets") if isinstance(raw.get("kb_snippets"), list) else None

    price_usd_raw = raw.get("price_usd", raw.get("price"))
    try:
      price_usd = float(price_usd_raw) if price_usd_raw is not None else 0.0
    except Exception:  # noqa: BLE001
      price_usd = 0.0

    price_cny_raw = raw.get("price_cny")
    try:
      price_cny = float(price_cny_raw) if price_cny_raw is not None else float(price_usd) * float(price_cny_rate)
    except Exception:  # noqa: BLE001
      price_cny = float(price_usd) * float(price_cny_rate)

    product_url = raw.get("product_url")
    image_url = raw.get("image_url")
    availability_raw = raw.get("availability")
    availability: List[str] = []
    if isinstance(availability_raw, list):
      availability = [str(x).strip() for x in availability_raw if str(x).strip()]
    elif isinstance(availability_raw, str) and availability_raw.strip():
      availability = [s.strip() for s in availability_raw.split(",") if s.strip()]
    if not availability:
      availability = ["Global"]

    if not brand or not name or not ingredients:
      print(f"Skipping JSON row {i}: missing brand/name/ingredients_text")
      continue

    skus.append(
      InputSku(
        brand=brand,
        name=name,
        ingredients_text=ingredients,
        price_usd=price_usd,
        price_cny=price_cny,
        availability=availability,
        product_url=str(product_url).strip() if product_url else None,
        image_url=str(image_url).strip() if image_url else None,
        expert_knowledge=expert_raw,
        kb_snippets=kb_snippets_raw,
      )
    )

    if limit is not None and len(skus) >= limit:
      break

  return skus


def demo_skus() -> List[InputSku]:
  return [
    InputSku(
      brand="Tom Ford",
      name="Research Serum Concentrate",
      price_usd=350.0,
      price_cny=2800.0,
      product_url="https://www.tomfordbeauty.com/...",
      ingredients_text="Water, Caffeine, Theobroma Cacao, Glycolic Acid, Alcohol Denat",
      availability=["Global"],
    ),
    InputSku(
      brand="The Ordinary",
      name="Buffet + Copper Peptides 1%",
      price_usd=30.0,
      price_cny=240.0,
      ingredients_text="Water, Glycerin, Copper Tripeptide-1, Lactococcus Ferment",
      availability=["Global"],
    ),
    InputSku(
      brand="Helena Rubinstein",
      name="Re-Plasty Age Recovery Night Cream (Black Bandage)",
      price_usd=460.0,
      price_cny=3900.0,
      ingredients_text="Water, Glycerin, Shea Butter, Dimethicone, Madecassoside, Fragrance",
      product_url="https://www.helenarubinstein.com/...",
      availability=["Global"],
    ),
  ]


def ingest_one(
  *,
  db: Optional[AuroraDb],
  client: GeminiClient,
  llm_model: str,
  embedding_model: str,
  social_provider: str,
  social_model: str,
  openai_api_key: Optional[str],
  openai_api_base_url: str,
  sku: InputSku,
  overwrite: bool,
  enable_embedding: bool,
  dry_run: bool,
) -> None:
  print(f"🧪 Processing: {sku.brand} - {sku.name} ...")

  if dry_run:
    vectors = get_vectors_from_llm(client, model=llm_model, brand=sku.brand, name=sku.name, ingredients=sku.ingredients_text)

    embedding: Optional[List[float]] = None
    if enable_embedding:
      embedding = get_embedding(client, model=embedding_model, text=sku.ingredients_text)

    social_payload: Dict[str, Any] = {}
    if social_provider == "openai":
      if not openai_api_key:
        raise RuntimeError("OPENAI_API_KEY is required when --social-provider=openai")
      print(f"   ...Simulating social stats for {sku.name}")
      social_payload = get_social_simulation_openai(
        brand=sku.brand,
        name=sku.name,
        ingredients_text=(sku.ingredients_text[:2000] if sku.ingredients_text else None),
        api_key=openai_api_key,
        model=social_model,
        api_base_url=openai_api_base_url,
      )
    elif social_provider == "llm":
      ss = vectors.get("social_stats") or {}
      if isinstance(ss, dict):
        social_payload = {
          "red_score": _clamp_int(ss.get("red_score") or ss.get("redScore"), min_value=0, max_value=100),
          "reddit_score": _clamp_int(ss.get("reddit_score") or ss.get("redditScore"), min_value=0, max_value=100),
          "burn_rate": _clamp_float(ss.get("burn_rate") or ss.get("burnRate"), min_value=0.0, max_value=1.0),
          "top_keywords": _coerce_keywords(ss.get("top_keywords") or ss.get("topKeywords")),
          "_raw": ss,
        }
    elif social_provider == "none":
      social_payload = {"red_score": 0, "reddit_score": 0, "burn_rate": 0.0, "top_keywords": []}

    print(
      json.dumps(
        {"product": {"brand": sku.brand, "name": sku.name}, "vectors": vectors, "social": social_payload, "expert_knowledge": sku.expert_knowledge},
        ensure_ascii=False,
      )
    )
    return

  if db is None:
    raise RuntimeError("Internal error: db is required unless --dry-run is set.")

  def _upsert_expert_knowledge(*, product_id: str) -> int:
    ek = sku.expert_knowledge
    if not isinstance(ek, dict) or not ek:
      return 0

    upserted = 0
    for key, value in ek.items():
      content = str(value).strip() if value is not None else ""
      if len(content) < 2:
        continue
      db.upsert_kb_snippet(
        product_id=product_id,
        source_sheet="expert_knowledge",
        field=canonicalize_kb_field(str(key)),
        content=content,
        metadata={
          "source": "expert_knowledge",
          "brand": sku.brand,
          "name": sku.name,
          # Provide a stable ontology key for downstream RAG bucketing.
          "canonical_key": infer_kb_canonical_key(str(key)) or "notes",
          "canonical_key_source": "expert_fields_v1",
        },
      )
      upserted += 1
    return upserted

  def _upsert_ingredient_derived_kb(*, product_id: str) -> int:
    ing = (sku.ingredients_text or "").strip()
    if len(ing) < 5:
      return 0

    ek = sku.expert_knowledge if isinstance(sku.expert_knowledge, dict) else {}
    expert_key_actives = ""
    if isinstance(ek, dict):
      expert_key_actives = str(ek.get("key_actives") or ek.get("key_actives_summary") or "").strip()

    key_actives = infer_key_actives_from_ingredients(ingredients_text=ing, expert_key_actives=expert_key_actives)
    risk_flags = infer_risk_flags_from_ingredients(product_name=f"{sku.brand} {sku.name}".strip(), ingredients_text=ing)

    upserted = 0

    if key_actives:
      db.upsert_kb_snippet(
        product_id=product_id,
        # "zz_" keeps these derived snippets after human notes in deterministic ordering.
        source_sheet="zz_ingredients_derived",
        field="key_actives",
        content=" | ".join(key_actives),
        metadata={
          "source": "ingredients_rules_v1",
          "canonical_key": "key_actives",
          "canonical_key_source": "ingredients_rules_v1",
          "brand": sku.brand,
          "name": sku.name,
        },
      )
      upserted += 1

    if risk_flags:
      db.upsert_kb_snippet(
        product_id=product_id,
        source_sheet="zz_ingredients_derived",
        field="sensitivity_flags",
        content=" | ".join(risk_flags),
        metadata={
          "source": "ingredients_rules_v1",
          "canonical_key": "sensitivity",
          "canonical_key_source": "ingredients_rules_v1",
          "brand": sku.brand,
          "name": sku.name,
        },
      )
      upserted += 1

    return upserted

  existing_id = db.find_product_id_loose(brand=sku.brand, name=sku.name)
  if existing_id and not overwrite:
    upserted = _upsert_expert_knowledge(product_id=str(existing_id))
    derived = _upsert_ingredient_derived_kb(product_id=str(existing_id))
    if upserted or derived:
      db.conn.commit()
      if upserted:
        print(f"📝 Expert knowledge upserted: {upserted}")
      if derived:
        print(f"🧩 Ingredient-derived KB upserted: {derived}")
    print(f"↩️  Skipped (exists): {sku.brand} - {sku.name}")
    return

  vectors = get_vectors_from_llm(client, model=llm_model, brand=sku.brand, name=sku.name, ingredients=sku.ingredients_text)
  embedding: Optional[List[float]] = None
  if enable_embedding:
    embedding = get_embedding(client, model=embedding_model, text=sku.ingredients_text)

  social_payload: Dict[str, Any] = {}
  if social_provider == "openai":
    if not openai_api_key:
      raise RuntimeError("OPENAI_API_KEY is required when --social-provider=openai")
    print(f"   ...Simulating social stats for {sku.name}")
    social_payload = get_social_simulation_openai(
      brand=sku.brand,
      name=sku.name,
      ingredients_text=(sku.ingredients_text[:2000] if sku.ingredients_text else None),
      api_key=openai_api_key,
      model=social_model,
      api_base_url=openai_api_base_url,
    )
  elif social_provider == "llm":
    ss = vectors.get("social_stats") or {}
    if isinstance(ss, dict):
      social_payload = {
        "red_score": _clamp_int(ss.get("red_score") or ss.get("redScore"), min_value=0, max_value=100),
        "reddit_score": _clamp_int(ss.get("reddit_score") or ss.get("redditScore"), min_value=0, max_value=100),
        "burn_rate": _clamp_float(ss.get("burn_rate") or ss.get("burnRate"), min_value=0.0, max_value=1.0),
        "top_keywords": _coerce_keywords(ss.get("top_keywords") or ss.get("topKeywords")),
        "_raw": ss,
      }
  elif social_provider == "none":
    social_payload = {"red_score": 0, "reddit_score": 0, "burn_rate": 0.0, "top_keywords": []}

  if existing_id and overwrite:
    product_id = existing_id
    # Keep product_id stable; replace dependent tables.
    db.upsert_product(sku, product_id=product_id)
    db.delete_vectors_for_product(product_id)
    db.delete_ingredients_for_product(product_id)
    db.delete_social_stats_for_product(product_id)
  else:
    product_id = str(uuid.uuid4())
  vector_id = str(uuid.uuid4())
  ingredient_id = str(uuid.uuid4())
  social_id = str(uuid.uuid4())

  try:
    if not existing_id:
      db.insert_product(sku, product_id=product_id)
    # Keep alias table warm for better anchor resolution in chat (best-effort).
    db.upsert_default_aliases_for_product(product_id=str(product_id), brand=sku.brand, name=sku.name)
    db.insert_vectors(
      product_id=product_id,
      vector_id=vector_id,
      mechanism=vectors["mechanism"],
      experience=vectors["experience_prediction"],
      risk_flags=vectors["risk_flags"],
      embedding=embedding,
    )
    db.insert_ingredients(product_id=product_id, ingredient_id=ingredient_id, full_list=parse_ingredients_list(sku.ingredients_text))
    db.insert_social_stats(
      product_id=product_id,
      social_id=social_id,
      red_score=int(social_payload.get("red_score", 0)),
      reddit_score=int(social_payload.get("reddit_score", 0)),
      burn_rate=float(social_payload.get("burn_rate", 0.0)),
      top_keywords=list(social_payload.get("top_keywords", []) if isinstance(social_payload.get("top_keywords", []), list) else []),
    )
    _upsert_expert_knowledge(product_id=str(product_id))
    _upsert_ingredient_derived_kb(product_id=str(product_id))
    db.conn.commit()
    print(f"✅ Ingested: {sku.brand} - {sku.name}")
  except BaseException:  # noqa: BLE001
    db.conn.rollback()
    raise


def main() -> None:
  parser = argparse.ArgumentParser(description="Aurora vectorization + embedding ETL (Excel → Gemini → Railway Postgres)")
  parser.add_argument("--demo", action="store_true", help="Ingest 3 demo products (Tom Ford / The Ordinary / HR).")
  parser.add_argument("--input", type=str, help="Path to Excel (.xlsx) file.")
  parser.add_argument("--input-json", type=str, help="Path to JSON file (list of SKUs) to ingest.")
  parser.add_argument("--sheet", type=str, default=None, help="Excel sheet name (default: active sheet).")
  parser.add_argument("--all-sheets", action="store_true", help="Ingest from all sheets in the workbook (deduped).")
  parser.add_argument("--list-sheets", action="store_true", help="List workbook sheet names and headers (requires --input).")
  parser.add_argument("--limit", type=int, default=None, help="Max rows to ingest (for testing).")
  parser.add_argument("--list-models", action="store_true", help="List available Gemini models and exit.")
  parser.add_argument("--ingest-kb", action="store_true", help="Also ingest non-ingredient notes into product_kb_snippets.")
  parser.add_argument("--kb-only", action="store_true", help="Only upsert KB snippets from the workbook (no LLM calls). Requires --input.")
  parser.add_argument(
    "--kb-bootstrap-products",
    action="store_true",
    help="In --kb-only mode: create missing Product rows (no vectors/ingredients) so KB snippets can attach even when INCI is missing.",
  )
  parser.add_argument("--aliases-only", action="store_true", help="Only upsert default Product Aliases from existing products (no LLM calls).")

  parser.add_argument("--col-brand", type=str, default="brand")
  parser.add_argument("--col-name", type=str, default="name")
  parser.add_argument("--col-ingredients", type=str, default="ingredients")
  parser.add_argument("--col-price-usd", type=str, default="price_usd")
  parser.add_argument("--col-price", type=str, default=None, help="Fallback price column if price_usd not present.")
  parser.add_argument("--price-cny-rate", type=float, default=7.2)
  parser.add_argument("--col-product-url", type=str, default=None)
  parser.add_argument("--col-image-url", type=str, default=None)

  parser.add_argument("--llm-model", type=str, default="gemini-2.5-flash")
  parser.add_argument("--embedding-model", type=str, default="gemini-embedding-001")
  parser.add_argument(
    "--social-provider",
    type=str,
    default=None,
    help="Social stats source: openai|llm|none (default: openai if OPENAI_API_KEY set, else llm).",
  )
  parser.add_argument("--social-model", type=str, default="gpt-4o", help="OpenAI model for social simulation (default: gpt-4o).")
  parser.add_argument("--no-embedding", action="store_true")
  parser.add_argument("--overwrite", action="store_true", help="Overwrite existing rows by (brand,name).")
  parser.add_argument("--dry-run", action="store_true", help="Call Gemini but do not write to DB.")

  args = parser.parse_args()

  load_dotenv()

  if args.list_sheets:
    if not args.input:
      raise SystemExit("--list-sheets requires --input /path/to.xlsx")
    wb = load_workbook(args.input, read_only=True, data_only=True)
    print("📄 Workbook sheets:")
    for sheet_name in wb.sheetnames:
      ws = wb[sheet_name]
      rows = ws.iter_rows(values_only=True)
      header = next(rows, None)
      headers = [str(h).strip() if h is not None else "" for h in (header or [])]
      print(f"- {sheet_name}: {headers}")
    return

  ran_special_mode = False

  if args.aliases_only:
    ran_special_mode = True
    database_url = sanitize_database_url_for_psycopg2(resolve_env_templates(_require_env("DATABASE_URL")))
    with AuroraDb(database_url) as db:
      rows = db.load_all_products_basic()
      upserted = 0
      for product_id, brand, name in rows:
        db.upsert_default_aliases_for_product(product_id=product_id, brand=brand, name=name)
        upserted += 1
      db.conn.commit()
      print(f"🔤 Product aliases upserted for {upserted} products.")

  if args.kb_only:
    ran_special_mode = True
    if not args.input:
      raise SystemExit("--kb-only requires --input /path/to.xlsx")

    database_url = sanitize_database_url_for_psycopg2(resolve_env_templates(_require_env("DATABASE_URL")))
    with AuroraDb(database_url) as db:
      snippets = extract_kb_snippets_from_workbook(path=args.input)
      if snippets:
        product_ids = db.load_all_product_ids()
        product_ids_norm = db.load_all_product_ids_normalized()
        product_ids_full_norm: Dict[str, str] = {}
        for (b, n), pid in product_ids.items():
          full_key = normalize_match_key(f"{b} {n}")
          if full_key:
            product_ids_full_norm.setdefault(full_key, pid)

        # Optional: bootstrap missing products so KB snippets can attach even when ingredient lists are absent.
        if args.kb_bootstrap_products:
          missing_products: Dict[str, int] = {}
          for snip in snippets:
            pid = product_ids.get((snip.brand, snip.name))
            if not pid:
              pid = product_ids_norm.get((normalize_match_key(snip.brand), normalize_match_key(snip.name)))
            if not pid:
              full_name = str(snip.metadata.get("product_full_name") or f"{snip.brand} {snip.name}").strip()
              pid = product_ids_full_norm.get(normalize_match_key(full_name))
            if not pid:
              full_name = str(snip.metadata.get("product_full_name") or f"{snip.brand} {snip.name}").strip()
              missing_products[full_name] = missing_products.get(full_name, 0) + 1

          if missing_products:
            created = 0
            for full_name, _count in sorted(missing_products.items(), key=lambda x: x[1], reverse=True):
              brand, name = safe_split_brand_and_name(full_name)
              pid = db.ensure_product_stub(brand=brand, name=name, availability=["Global"])
              # Update local maps so the upcoming upsert pass can attach immediately.
              product_ids[(brand, name)] = pid
              product_ids_norm[(normalize_match_key(brand), normalize_match_key(name))] = pid
              product_ids_full_norm[normalize_match_key(full_name)] = pid
              created += 1
            db.conn.commit()
            print(f"🧱 Bootstrapped missing products: {created}")

        upserted = 0
        skipped = 0
        missing_counts: Dict[str, int] = {}
        for snip in snippets:
          pid = product_ids.get((snip.brand, snip.name))
          if not pid:
            pid = product_ids_norm.get((normalize_match_key(snip.brand), normalize_match_key(snip.name)))
          if not pid:
            full_name = str(snip.metadata.get("product_full_name") or f"{snip.brand} {snip.name}").strip()
            pid = product_ids_full_norm.get(normalize_match_key(full_name))
          if not pid:
            skipped += 1
            full_name = str(snip.metadata.get("product_full_name") or f"{snip.brand} {snip.name}").strip()
            missing_counts[full_name] = missing_counts.get(full_name, 0) + 1
            continue
          db.upsert_kb_snippet(
            product_id=pid,
            source_sheet=snip.source_sheet,
            field=snip.field,
            content=snip.content,
            metadata=snip.metadata,
          )
          upserted += 1
        db.conn.commit()
        print(f"📚 KB snippets upserted: {upserted} (skipped missing products: {skipped})")
        if missing_counts:
          print("⚠️  Missing products (top 20 by missing snippet count):")
          for name, count in sorted(missing_counts.items(), key=lambda x: x[1], reverse=True)[:20]:
            print(f"  - {name}  missing_snippets={count}")
      else:
        print("📚 No KB snippets found in workbook.")

  # If the user requested one of the "special modes", do not proceed to normal ingestion.
  if ran_special_mode:
    return

  api_key = _require_any_env(["GEMINI_API_KEY", "GOOGLE_API_KEY"])
  api_base_url = (os.getenv("GEMINI_API_BASE_URL") or DEFAULT_GEMINI_API_BASE_URL).strip()
  openai_api_key = (os.getenv("OPENAI_API_KEY") or "").strip()
  openai_api_base_url = (os.getenv("OPENAI_API_BASE_URL") or "https://api.openai.com/v1").strip()

  social_provider = (args.social_provider or "").strip().lower() or ("openai" if openai_api_key else "llm")
  if social_provider not in ("openai", "llm", "none"):
    raise SystemExit("--social-provider must be one of: openai|llm|none")
  if social_provider == "openai" and not openai_api_key:
    raise RuntimeError("OPENAI_API_KEY is required when --social-provider=openai (or when OPENAI_API_KEY is set default to openai).")

  gemini_client = GeminiClient(api_key=api_key, api_base_url=api_base_url)

  if args.list_models:
    models = gemini_client.list_models()
    # Keep output compact & copyable
    for m in models:
      name = m.get("name")
      methods = m.get("supportedGenerationMethods") or []
      short = GeminiClient.normalize_model_name(str(name or ""))
      print(f"{short}  full={name}  methods={methods}")
    return

  if not args.demo and not args.input and not args.input_json:
    raise SystemExit("Provide --demo, --input /path/to.xlsx, or --input-json /path/to.json (or --list-models)")

  if args.demo:
    skus = demo_skus()
  elif args.input_json:
    skus = load_input_skus_from_json(path=args.input_json, price_cny_rate=float(args.price_cny_rate), limit=args.limit)
  else:
    if args.all_sheets:
      if args.sheet:
        print("ℹ️  --all-sheets is set; ignoring --sheet.")
      skus = load_input_skus_from_workbook_all_sheets(
        path=args.input,
        col_brand=args.col_brand,
        col_name=args.col_name,
        col_ingredients=args.col_ingredients,
        col_price_usd=args.col_price_usd,
        col_price=args.col_price,
        price_cny_rate=float(args.price_cny_rate),
        col_product_url=args.col_product_url,
        col_image_url=args.col_image_url,
        limit=args.limit,
      )
    else:
      skus = load_input_skus_from_excel(
        path=args.input,
        sheet=args.sheet,
        col_brand=args.col_brand,
        col_name=args.col_name,
        col_ingredients=args.col_ingredients,
        col_price_usd=args.col_price_usd,
        col_price=args.col_price,
        price_cny_rate=float(args.price_cny_rate),
        col_product_url=args.col_product_url,
        col_image_url=args.col_image_url,
        limit=args.limit,
      )

  if not skus:
    print("No rows found to ingest.")
    return

  if args.dry_run:
    for sku in skus:
      ingest_one(
        db=None,
        client=gemini_client,
        llm_model=args.llm_model,
        embedding_model=args.embedding_model,
        social_provider=social_provider,
        social_model=args.social_model,
        openai_api_key=openai_api_key if social_provider == "openai" else None,
        openai_api_base_url=openai_api_base_url,
        sku=sku,
        overwrite=bool(args.overwrite),
        enable_embedding=not bool(args.no_embedding),
        dry_run=True,
      )
    return

  database_url = sanitize_database_url_for_psycopg2(resolve_env_templates(_require_env("DATABASE_URL")))
  with AuroraDb(database_url) as db:
    for sku in skus:
      ingest_one(
        db=db,
        client=gemini_client,
        llm_model=args.llm_model,
        embedding_model=args.embedding_model,
        social_provider=social_provider,
        social_model=args.social_model,
        openai_api_key=openai_api_key if social_provider == "openai" else None,
        openai_api_base_url=openai_api_base_url,
        sku=sku,
        overwrite=bool(args.overwrite),
        enable_embedding=not bool(args.no_embedding),
        dry_run=bool(args.dry_run),
      )

    if args.ingest_kb and args.input:
      snippets = extract_kb_snippets_from_workbook(path=args.input)
      if snippets:
        product_ids = db.load_all_product_ids()
        product_ids_norm = db.load_all_product_ids_normalized()
        product_ids_full_norm: Dict[str, str] = {}
        for (b, n), pid in product_ids.items():
          full_key = normalize_match_key(f"{b} {n}")
          if full_key:
            product_ids_full_norm.setdefault(full_key, pid)
        upserted = 0
        skipped = 0
        for snip in snippets:
          pid = product_ids.get((snip.brand, snip.name))
          if not pid:
            pid = product_ids_norm.get((normalize_match_key(snip.brand), normalize_match_key(snip.name)))
          if not pid:
            full_name = str(snip.metadata.get("product_full_name") or f"{snip.brand} {snip.name}").strip()
            pid = product_ids_full_norm.get(normalize_match_key(full_name))
          if not pid:
            skipped += 1
            continue
          db.upsert_kb_snippet(
            product_id=pid,
            source_sheet=snip.source_sheet,
            field=snip.field,
            content=snip.content,
            metadata=snip.metadata,
          )
          upserted += 1
        db.conn.commit()
        print(f"📚 KB snippets upserted: {upserted} (skipped missing products: {skipped})")
      else:
        print("📚 No KB snippets found in workbook.")

    if args.ingest_kb and args.input_json:
      product_ids = db.load_all_product_ids()
      product_ids_norm = db.load_all_product_ids_normalized()
      product_ids_full_norm: Dict[str, str] = {}
      for (b, n), pid in product_ids.items():
        full_key = normalize_match_key(f"{b} {n}")
        if full_key:
          product_ids_full_norm.setdefault(full_key, pid)

      upserted = 0
      skipped = 0
      for sku in skus:
        snippets_raw = sku.kb_snippets if isinstance(sku.kb_snippets, list) else []
        if not snippets_raw:
          continue

        pid = product_ids.get((sku.brand, sku.name))
        if not pid:
          pid = product_ids_norm.get((normalize_match_key(sku.brand), normalize_match_key(sku.name)))
        if not pid:
          pid = product_ids_full_norm.get(normalize_match_key(f"{sku.brand} {sku.name}"))
        if not pid:
          skipped += len(snippets_raw)
          continue

        for raw in snippets_raw:
          if not isinstance(raw, dict):
            continue
          source_sheet = str(raw.get("source_sheet") or raw.get("sourceSheet") or "json").strip() or "json"
          field = str(raw.get("field") or raw.get("canonical_key") or raw.get("canonicalKey") or "notes").strip() or "notes"
          content = str(raw.get("content") or "").strip()
          if not content:
            continue

          metadata = dict(raw)
          metadata.setdefault("ingested_from", "json")
          metadata.setdefault("product_full_name", f"{sku.brand} {sku.name}".strip())
          db.upsert_kb_snippet(
            product_id=pid,
            source_sheet=source_sheet,
            field=field,
            content=content,
            metadata=metadata,
          )
          upserted += 1

      db.conn.commit()
      print(f"📚 KB snippets upserted from JSON: {upserted} (skipped missing products: {skipped})")


if __name__ == "__main__":
  main()
